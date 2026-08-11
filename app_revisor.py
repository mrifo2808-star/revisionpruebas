"""
Revisor de Hojas de Respuestas — App Streamlit
- Pauta dinámica (N preguntas configurables)
- Carga de fotos desde celular o PC
- Edición de datos del alumno y corrección de respuestas dudosas
- Exporta Excel con resumen, detalle y estadísticas
"""

import io
import json
import re
import base64
import hashlib
from collections import Counter
import qrcode
import pandas as pd
import streamlit as st
import anthropic
import registro_sheets
from PIL import Image, ImageEnhance, ImageOps
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Motor OMR (opcional): si opencv/numpy no están disponibles en el entorno de
# despliegue, la app sigue funcionando 100% igual que antes con el flujo solo-IA
# — el modo OMR simplemente no aparece como opción en el sidebar. El motor va
# inline en este mismo archivo (no en un módulo omr.py aparte) a propósito:
# tener dos archivos abrió la puerta, más de una vez en producción, a que
# Streamlit Cloud sirviera una combinación inconsistente de ambos durante un
# redeploy (un archivo ya actualizado, el otro todavía en la versión previa),
# produciendo errores confusos que dependían de qué mitad del código corriera
# en ese momento. Con un solo archivo esa clase de bug deja de ser posible.
try:
    import numpy as np
    import cv2
    OMR_DISPONIBLE = True
except Exception:
    OMR_DISPONIBLE = False

# ═══════════════════════════════════════════════════════════════════════
# MOTOR OMR (Optical Mark Recognition) — plantilla "PLANTILLA DE HOJA DE
# RESPUESTAS". Detecta automáticamente qué alternativa (A-E) marcó el
# estudiante en cada fila de la tabla RESPUESTAS. Este motor es la ÚNICA
# fuente de las respuestas: Claude Vision nunca lee ni confirma burbujas (la
# app la usa solo para transcribir nombre/RUT de la cabecera, tarea aparte).
# Las preguntas que el motor no logra determinar con confianza quedan como
# "dudosas" para revisión manual en la app — ninguna IA adivina ni corrige
# una lectura de burbujas.
#
# Diseño: en vez de coordenadas de píxel fijas (frágil ante fotos de celular
# con ángulos/distancias distintas), todo se deriva de la propia imagen:
#   1. localizar el bloque RESPUESTAS (umbral adaptativo + contornos) solo si
#      la imagen es la hoja completa; si ya viene recortada se usa entera;
#   2. si se localizó, enderezarlo con una transformación de perspectiva;
#   3. ubicar la barra de encabezado y las bandas de columnas (1-4) por
#      proyección de tinta, probando varios umbrales de sensibilidad; si
#      ninguno distingue suficientes huecos reales entre columnas (foto
#      borrosa/comprimida) Y la región ya fue validada como la tabla
#      completa, se reparte el ancho de contenido en partes iguales en vez
#      de fallar — nunca se reescala ni comprime la imagen, solo cambia
#      dónde se traza el límite de cada columna. En un recorte sin validar
#      (subido directo por el usuario), ante la duda se prefiere fallar
#      claro antes que fabricar una columna que no existe;
#   4. ajustar, POR BANDA POR SEPARADO, una grilla de 20 filas x 5 columnas
#      (Hough circles + kmeans 1D) — las filas no se comparten entre bandas
#      a propósito, para que un corrimiento en una columna no contamine a
#      las demás (encontrado con verificación manual sobre fotos reales);
#   5. medir cuánto más oscura está cada burbuja que el papel en blanco,
#      comparando SIEMPRE dentro de la misma fila, con baseline/peak
#      calculados POR BANDA (no globales) para tolerar iluminación despareja.
#
# Detección a escala de referencia, medición a resolución completa: los pasos
# 1-4 (encontrar dónde están la tabla, las columnas, las filas y las burbujas)
# se calculan siempre sobre una copia reescalada a un ancho de referencia fijo
# (ver _omr_escalar_para_deteccion) -- NUNCA sobre la foto original completa --
# y esas coordenadas se escalan de vuelta antes del paso 5. Se encontró de
# forma directa que sin este paso, la misma foto real que da 100% de exactitud
# a la resolución con la que se calibró Hough/blur cae a ~21% con solo 4/80
# preguntas resueltas si se la reescala 4x (una foto de celular moderna
# fácilmente tiene 4-6x más resolución que esa referencia): el blur de kernel
# fijo y los umbrales de Hough dejan de ser apropiados para círculos mucho más
# grandes en píxeles. La foto original en sí NUNCA se reduce -- el paso 5
# (medir oscuridad, recortar preguntas dudosas) siempre usa la resolución
# completa; solo la búsqueda de geometría usa una copia más chica.
# ═══════════════════════════════════════════════════════════════════════

OMR_THRESHOLDS = {
    "MIN_MARK_SCORE": 0.25,          # confianza mínima para considerar que SÍ hay una marca real
    "AMBIGUOUS_MARGIN": 0.15,        # margen mínimo entre 1ª y 2ª alternativa para no ser ambiguo/empate
    "HIGH_CONFIDENCE_MARGIN": 0.30,  # margen a partir del cual se usa directo, sin marcar dudosa
    "DOUBLE_MARK_THRESHOLD": 0.55,   # si dos alternativas superan esto, se considera "doble marca"
    "BLANK_REVIEW_MARGIN": 0.12,     # "sin_marca" por encima de esto = posible marca débil, queda dudosa
    "MIN_GEOMETRY_CONFIDENCE": 0.5,  # por debajo de esto, la banda no tiene evidencia real de ser
                                      # una grilla de burbujas (pudo caer sobre texto/margen vecino) --
                                      # ninguna pregunta de esa banda puede darse como confiable
                                      # (estado GEOMETRY_ERROR)
    "WARNING_GEOMETRY_CONFIDENCE": 0.8,  # entre este valor y MIN_GEOMETRY_CONFIDENCE: geometría
                                          # aceptable pero con alguna señal débil (spacing irregular,
                                          # radio cerca del límite) -- estado GEOMETRY_WARNING, la
                                          # pregunta puede seguir siendo confiable pero queda marcada
                                          # para que un diagnóstico visual la destaque
}
OMR_N_FILAS_POR_BLOQUE = 20  # fijo por diseño de la plantilla impresa (igual que usa el prompt de Claude)

# Feature flag: el NUMBER LATTICE (segundo sensor geométrico sobre los
# números de fila impresos, ver omr_ajustar_number_lattice) se calcula y se
# expone SIEMPRE en diagnóstico, pero su veto sobre geometry_confidence
# arranca OFF por defecto. Motivo medido, no preventivo: en la foto de
# calibración (el fixture con ground truth más confiable que hay) el
# crosscheck disparó un falso positivo -- el dígito "1" de la fila 1 es
# demasiado delgado para producir tinta detectable a la resolución de
# detección (~480px de ancho de tabla), así que el lattice de números quedó
# corrido una fila completa contra el de burbujas (que sí es correcto,
# verificado contra ground truth) y hubiera tirado geometry_confidence de
# una banda sana a GEOMETRY_ERROR. Mismo criterio que AI_ROW_ANCHOR_CHECK
# (plan item 20): no se habilita un veto adicional sin evidencia de que
# aporta más de lo que arriesga -- requiere muestrear la franja de números a
# mayor resolución antes de poder usarse como gate real.
NUMBER_LATTICE_CROSSCHECK_HABILITADO = False
OMR_LETRAS = ["A", "B", "C", "D", "E"]

# Perfil de la plantilla física impresa. Existe para que el N máximo de
# preguntas configurable en la UI (más abajo) no pueda superar lo que la
# hoja física puede mostrar -- antes la UI aceptaba hasta 120 preguntas
# aunque la plantilla real solo tiene 4 bloques x 20 filas = 80, una
# contradicción real entre estructura física e input arbitrario. Un formato
# de hoja distinto en el futuro sería OTRO perfil, no un N más grande sobre
# este mismo.
TEMPLATE_PROFILE = {
    "id": "PAES_80_V1",
    "n_bloques_max": 4,
    "n_filas_por_bloque": OMR_N_FILAS_POR_BLOQUE,
    "n_alternativas": len(OMR_LETRAS),
    "n_max": 4 * OMR_N_FILAS_POR_BLOQUE,  # 80 -- techo físico real de esta plantilla
}
OMR_MAX_BANDAS = TEMPLATE_PROFILE["n_bloques_max"]


class OMRError(Exception):
    """Fallo irrecuperable de una etapa del pipeline OMR — la hoja queda sin
    resolver para completar a mano (nunca se reemplaza con una lectura de IA)."""


def _omr_encontrar_barra_encabezado(gray):
    """
    Devuelve el borde inferior (y) de la barra de encabezado oscura y sólida
    en la parte superior de la imagen, o 0 si no se encuentra ninguna. Misma
    lógica que usa omr_detectar_header_y_bandas() para saltar la barra
    "RESPUESTAS" antes de leer la tabla -- se comparte acá para no duplicarla.
    """
    h = gray.shape[0]
    row_mean = gray.mean(axis=1)
    dark_rows = row_mean < (row_mean.mean() - row_mean.std() * 0.5)
    header_bottom = 0
    in_run, run_start = False, 0
    for y in range(min(h, int(h * 0.25))):
        if dark_rows[y] and not in_run:
            in_run, run_start = True, y
        elif not dark_rows[y] and in_run:
            in_run = False
            if (y - run_start) > h * 0.015:
                header_bottom = y
    return header_bottom


def _omr_contar_picos_locales(hist, min_valor):
    """Cuenta máximos locales en un histograma 1D (bins consecutivos): un bin
    es un pico si es >= a ambos vecinos y supera `min_valor`. A diferencia de
    "contar corridas de bins por encima de un piso", esto SÍ separa filas
    reales adyacentes aunque el valle entre ellas no baje hasta un piso fijo
    -- una fila real de burbujas contra otra puede compartir un valle de
    conteo moderado (no cero) por el solapamiento de detecciones Hough
    vecinas, y una corrida-por-piso las fusiona en una sola "fila" (bug real:
    una franja con 6 filas reales de burbujas medía solo 1-3 corridas con esa
    lógica, según el piso elegido, mientras que con máximos locales mide 5-6,
    coincidiendo con las filas reales visibles en la foto)."""
    n = len(hist)
    picos = []
    for i in range(n):
        if hist[i] < min_valor:
            continue
        if (i == 0 or hist[i] >= hist[i - 1]) and (i == n - 1 or hist[i] >= hist[i + 1]):
            picos.append(i)
    fusionados = []
    for p in picos:
        if fusionados and p - fusionados[-1] <= 1:
            continue
        fusionados.append(p)
    return len(fusionados)


def _omr_header_seguido_de_grilla(warp_gray, header_bottom, min_filas_densas=4, min_conteo_pico=3):
    """
    Verifica que la franja de contenido INMEDIATAMENTE debajo del header
    detectado sea realmente el cuerpo de burbujas de RESPUESTAS, no otra
    sección de la hoja que también tenga su propia barra oscura y sólida
    real. `_omr_encontrar_barra_encabezado` solo confirma que existe ALGUNA
    barra -- no distingue "RESPUESTAS" de "IDENTIFICACIÓN DEL ESTUDIANTE"
    (sin OCR no puede). Encontrado con smoke test real: cuando el contorno
    del candidato fusiona/aísla la sección de identificación en vez de
    RESPUESTAS, la barra "IDENTIFICACIÓN DEL ESTUDIANTE" -- real, sólida --
    pasaba el check en vez de la barra "RESPUESTAS" real, y el resto del
    pipeline terminaba leyendo casillas de nombre/folleto como si fueran
    burbujas de respuesta.

    Ni contar círculos Hough en la franja, ni exigir columnas parejas, ni
    clusterizar filas por simple espaciado de gaps alcanza por sí solo:
    letras manuscritas y bordes de casillas de texto también producen
    "círculos" Hough y hasta clusters de columna/fila con spacing parecido
    (probado sobre fotos reales) -- y un candidato de identificación con
    varias filas de casillero (apellido paterno/materno/nombres) puede
    incluso producir 1-2 "filas" de ruido agrupado, aunque nunca tantas como
    una tabla real. La señal que sí separa ambos casos con margen amplio
    (~5 filas reales vs. ~2 como máximo de ruido, medido sobre fotos reales)
    es la DENSIDAD DE FILAS: una tabla RESPUESTAS empaqueta ~20 filas de
    burbujas en el espacio que una sola fila de un casillero de texto ocupa
    para sí sola. Se cuenta como histograma de densidad (no clustering por
    gaps, que el ruido disperso entre filas reales puede puentear sin dejar
    un hueco limpio) cuántas "filas" (bins consecutivos con >= min_conteo_pico
    círculos) hay dentro de una franja de altura fija relativa al ANCHO del
    candidato (más estable que relativa a su alto, que es justamente lo que
    puede estar inflado o recortado por una fusión/aislamiento de secciones)."""
    h_total, w_total = warp_gray.shape
    alto_franja = max(15, int(w_total * 0.18))
    y0, y1 = header_bottom, min(h_total, header_bottom + alto_franja)
    if y1 - y0 < 8:
        return False
    franja = warp_gray[y0:y1, :]
    franja_det, _ = _omr_escalar_para_deteccion(franja, OMR_ANCHO_REF_TABLA)
    blur = cv2.medianBlur(franja_det, 3)
    ww = franja_det.shape[1]
    r_guess = max(3, ww // 60)
    circles = cv2.HoughCircles(
        blur, cv2.HOUGH_GRADIENT, dp=1, minDist=max(4, r_guess),
        param1=60, param2=14, minRadius=max(2, int(r_guess * 0.5)),
        maxRadius=int(r_guess * 2.2),
    )
    if circles is None:
        return False
    ys = circles[0][:, 1]
    ancho_bin = max(2.0, r_guess * 0.5)
    bins = np.arange(0, franja_det.shape[0] + ancho_bin, ancho_bin)
    hist, _ = np.histogram(ys, bins=bins)
    return _omr_contar_picos_locales(hist, min_conteo_pico) >= min_filas_densas


def _omr_candidato_tiene_header(gray, rect):
    """
    Verifica que la parte superior del candidato tenga una barra de
    encabezado oscura y sólida real -- como la barra "RESPUESTAS" impresa en
    la plantilla -- E INMEDIATAMENTE debajo de ella haya evidencia real de
    burbujas (ver _omr_header_seguido_de_grilla). Distingue el bloque
    RESPUESTAS real de un contorno que, por poca luz o una sombra, terminó
    fusionado con contenido vecino en la misma hoja (p.ej. el texto "USO
    EXCLUSIVO PARA ENSAYOS DE PRUEBAS" impreso al lado de la tabla, o la
    sección "IDENTIFICACIÓN DEL ESTUDIANTE" fusionada por encima) -- ambos
    casos encontrados con fotos reales donde el bloque detectado terminaba
    leyendo texto/casillas vecinas en vez de burbujas de respuesta.
    """
    box = cv2.boxPoints(rect).astype("float32")
    src = _omr_ordenar_puntos(box)
    w_ = int(max(np.linalg.norm(src[1] - src[0]), np.linalg.norm(src[2] - src[3])))
    h_ = int(max(np.linalg.norm(src[3] - src[0]), np.linalg.norm(src[2] - src[1])))
    if w_ < 20 or h_ < 20:
        return False
    dst = np.array([[0, 0], [w_ - 1, 0], [w_ - 1, h_ - 1], [0, h_ - 1]], dtype="float32")
    M = cv2.getPerspectiveTransform(src, dst)
    warp = cv2.warpPerspective(gray, M, (w_, h_))
    header_bottom = _omr_encontrar_barra_encabezado(warp)
    if header_bottom <= 0:
        return False
    return _omr_header_seguido_de_grilla(warp, header_bottom)


def _omr_candidato_evidencia_grilla(gray, rect, min_circulos=80, min_filas_densas=15, min_conteo_pico=3):
    """
    Evidencia geométrica INDEPENDIENTE del header: ¿el candidato realmente
    contiene un patrón denso y repetitivo de burbujas (círculos chicos), o es
    una región sólida por alguna otra razón (bloque de texto denso, sombra,
    borde de página)? Puramente geométrica -- nunca lee el contenido del
    texto (sin OCR). Es la única base aceptable para usar igual un candidato
    que no tiene una barra de encabezado real: sin esto, "el candidato más
    grande" no demuestra nada sobre si ES la tabla RESPUESTAS.

    min_circulos=80 es un piso bajo a propósito (una tabla real de 1-4
    columnas x 20 filas x 5 alternativas tiene entre 100 y 400 burbujas) --
    solo busca descartar candidatos que casi no tienen círculos reales
    (texto, ruido de compresión), no exigir una cuenta exacta. PERO no
    alcanza por sí solo: la grilla de dígitos "CÉDULA DE IDENTIDAD"/
    "PASAPORTE" de esta misma plantilla también es un patrón denso y
    repetitivo de círculos reales (encontrado en smoke test real: 107
    círculos, sobre el piso de 80) -- es una grilla real, solo que la
    incorrecta. La distingue su NÚMERO DE FILAS: CÉDULA tiene ~10-11 filas
    (dígitos 0-9, +K), RESPUESTAS tiene hasta 20 por banda -- se exige
    encontrar al menos `min_filas_densas` filas densas de círculos en el
    candidato completo (mismo método de histograma de densidad que
    `_omr_header_seguido_de_grilla`, reutilizado acá sobre todo el
    candidato en vez de solo la franja bajo el header)."""
    box = cv2.boxPoints(rect).astype("float32")
    src = _omr_ordenar_puntos(box)
    w_ = int(max(np.linalg.norm(src[1] - src[0]), np.linalg.norm(src[2] - src[3])))
    h_ = int(max(np.linalg.norm(src[3] - src[0]), np.linalg.norm(src[2] - src[1])))
    if w_ < 20 or h_ < 20:
        return False
    dst = np.array([[0, 0], [w_ - 1, 0], [w_ - 1, h_ - 1], [0, h_ - 1]], dtype="float32")
    M = cv2.getPerspectiveTransform(src, dst)
    warp = cv2.warpPerspective(gray, M, (w_, h_))
    warp_det, _ = _omr_escalar_para_deteccion(warp, OMR_ANCHO_REF_TABLA)
    blur = cv2.medianBlur(warp_det, 3)
    ww = warp_det.shape[1]
    r_guess = max(3, ww // 60)  # orden de magnitud de una burbuja en una tabla de 1-4 columnas
    circles = cv2.HoughCircles(
        blur, cv2.HOUGH_GRADIENT, dp=1, minDist=max(4, r_guess),
        param1=60, param2=14, minRadius=max(2, int(r_guess * 0.5)),
        maxRadius=int(r_guess * 2.2),
    )
    if circles is None or circles.shape[1] < min_circulos:
        return False
    ys = circles[0][:, 1]
    ancho_bin = max(2.0, r_guess * 0.5)
    bins = np.arange(0, warp_det.shape[0] + ancho_bin, ancho_bin)
    hist, _ = np.histogram(ys, bins=bins)
    return _omr_contar_picos_locales(hist, min_conteo_pico) >= min_filas_densas


def omr_detectar_bloque_respuestas(img_bgr):
    """
    Devuelve los 4 puntos (rotados, en cualquier orden) del rectángulo que
    contiene la tabla RESPUESTAS dentro de una foto de la hoja completa, o
    None si no se encuentra con confianza suficiente (la hoja completa tiene
    además los bloques de identificación/cédula, más chicos y con menor
    fill_ratio que el bloque RESPUESTAS, el mayor recuadro "sólido" del diseño).
    """
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape
    th = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                cv2.THRESH_BINARY_INV, 35, 15)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (9, 9))
    # iterations=1 (antes 2): con 2 iteraciones el alcance efectivo del cierre
    # (~16-18px a esta escala de referencia) puentea el hueco real entre
    # secciones DISTINTAS de la hoja (medido en una foto real: 13-14px entre
    # "IDENTIFICACIÓN DEL ESTUDIANTE"/"N° DE FOLLETO"/"RESPUESTAS", contra
    # 2-6px de hueco DENTRO de una misma sección) -- fusionaba identificación
    # + folleto + RESPUESTAS en un solo contorno gigante, y el bloque
    # "detectado" terminaba siendo casi toda la página. 1 iteración sigue
    # cerrando los huecos internos de una sección (más chicos) sin puentear
    # el hueco real entre secciones distintas.
    closed = cv2.morphologyEx(th, cv2.MORPH_CLOSE, kernel, iterations=1)
    contours, _ = cv2.findContours(closed, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)

    candidatos = []
    page_area = h * w
    for c in contours:
        area = cv2.contourArea(c)
        if area < page_area * 0.03:
            continue
        rect = cv2.minAreaRect(c)
        (_, _), (rw, rh), _ = rect
        if rw < 1 or rh < 1:
            continue
        fill = area / (rw * rh)
        long_side, short_side = max(rw, rh), min(rw, rh)
        aspect = long_side / max(short_side, 1e-6)
        # el bloque RESPUESTAS es moderadamente apaisado (4 columnas una al
        # lado de otra); evita capturar la hoja entera o cajas chicas de
        # cédula/ID. Umbrales deliberadamente permisivos (calibrados sobre
        # pocas fotos reales): mejor aceptar de más que rechazar la tabla
        # real por una foto con encuadre distinto.
        #
        # fill>0.15 (antes 0.45): una tabla real de 20 filas x hasta 4
        # bandas tiene MUCHO espacio en blanco entre filas/columnas -- medido
        # en una foto real donde el bloque RESPUESTAS correcto, limpio y
        # bien aislado (confirmado visualmente) media fill=0.21, muy por
        # debajo del piso anterior. Ese piso alto no protegía contra nada
        # real (el bloque de texto denso sintético que sí debe rechazarse
        # mide fill=0.73, muy por encima) -- la protección real contra texto
        # viene de _omr_candidato_tiene_header / _omr_candidato_evidencia_grilla
        # (exigen barra de encabezado + densidad de filas de círculos reales),
        # no del fill ratio.
        if fill > 0.15 and 0.9 < aspect < 3.2 and area < page_area * 0.75:
            candidatos.append((area, rect))
    if not candidatos:
        return None
    candidatos.sort(key=lambda t: t[0], reverse=True)
    # Entre los candidatos, preferir el más grande que además tenga una barra
    # de encabezado real.
    for area, rect in candidatos:
        if _omr_candidato_tiene_header(gray, rect):
            return cv2.boxPoints(rect)
    # Ningún candidato tiene un header real. ANTES se usaba igual el más
    # grande ("mejor arriesgarse a leer de más que fallar") -- ese es
    # exactamente el mecanismo que produjo el bug real "USO EXCLUSIVO": un
    # candidato sin evidencia de ser la tabla se devolvía como si lo fuera
    # solo por tamaño. Regla actual: NO HEADER + NO EVIDENCIA DE GRILLA = NO
    # TABLA. Sin header, se exige evidencia geométrica independiente (patrón
    # denso de círculos reales, no el tamaño del rectángulo) antes de aceptar
    # igual un candidato -- nunca OCR, nunca buscar el texto literal.
    for area, rect in candidatos:
        if _omr_candidato_evidencia_grilla(gray, rect):
            return cv2.boxPoints(rect)
    return None


def _omr_ordenar_puntos(pts):
    pts = np.array(pts)
    s = pts.sum(axis=1)
    d = np.diff(pts, axis=1).flatten()
    tl = pts[np.argmin(s)]
    br = pts[np.argmax(s)]
    tr = pts[np.argmin(d)]
    bl = pts[np.argmax(d)]
    return np.array([tl, tr, br, bl], dtype="float32")


def omr_enderezar_region(img_bgr, quad):
    """Perspective-warp de un cuadrilátero (4 puntos, cualquier orden/rotación) a un rectángulo recto."""
    src = _omr_ordenar_puntos(quad)
    w_top = np.linalg.norm(src[1] - src[0])
    w_bot = np.linalg.norm(src[2] - src[3])
    h_left = np.linalg.norm(src[3] - src[0])
    h_right = np.linalg.norm(src[2] - src[1])
    out_w = int(max(w_top, w_bot))
    out_h = int(max(h_left, h_right))
    dst = np.array([[0, 0], [out_w - 1, 0], [out_w - 1, out_h - 1], [0, out_h - 1]], dtype="float32")
    M = cv2.getPerspectiveTransform(src, dst)
    return cv2.warpPerspective(img_bgr, M, (out_w, out_h))


def _omr_bandas_por_gaps(col_smooth, bw, umbral):
    mask = col_smooth > umbral
    bands = []
    in_band, start = False, 0
    for x in range(bw):
        if mask[x] and not in_band:
            in_band, start = True, x
        elif not mask[x] and in_band:
            in_band = False
            bands.append((start, x))
    if in_band:
        bands.append((start, bw))
    bands = [b for b in bands if (b[1] - b[0]) > bw * 0.04]
    return bands, mask


def omr_detectar_header_y_bandas(gray, max_bandas=OMR_MAX_BANDAS, bandas_esperadas=None, permitir_reparto_geometrico=True):
    """
    Devuelve (header_bottom, bandas, bandas_fabricadas) donde bandas es una
    lista de (x0,x1) — una por cada bloque de 20 preguntas visible en la
    imagen (1 a 4). header_bottom=0 si no se detecta una barra de encabezado
    sólida (imagen ya recortada que empieza directo en la fila 1).
    bandas_fabricadas=True si los límites de columna NO se pudieron
    distinguir por huecos reales de tinta y se repartió el ancho
    proporcionalmente en su lugar -- se propaga a geometry_source por banda
    en omr_ajustar_grilla (GEOMETRIC_BAND_FALLBACK) para que esa geometría
    nunca reciba la misma confianza que una detectada de verdad.

    `permitir_reparto_geometrico`: si los huecos reales entre columnas no se
    distinguen con ningún umbral, repartir el ancho de contenido en partes
    iguales asume que el contenido visible SÍ contiene las `bandas_esperadas`
    columnas completas. Razonable cuando la región ya fue validada como la
    tabla RESPUESTAS completa; peligroso en un recorte sin validar que
    genuinamente no muestre todas las columnas (fabricaría una columna
    inexistente y devolvería respuestas con confianza alta pero incorrectas).
    """
    h, w = gray.shape
    header_bottom = _omr_encontrar_barra_encabezado(gray)
    body = gray[header_bottom:h, :]
    bh, bw = body.shape

    col_ink = 255 - body.astype(np.float32).mean(axis=0)
    rng = col_ink.max() - col_ink.min()
    if rng < 8:
        raise OMRError("La imagen no tiene suficiente contraste para ubicar columnas "
                        "(¿es realmente una foto de la tabla RESPUESTAS?).")
    col_norm = (col_ink - col_ink.min()) / rng
    k = np.ones(5) / 5
    col_smooth = np.convolve(col_norm, k, mode="same")

    objetivo = bandas_esperadas or max_bandas
    bands, mask = [], col_smooth > 0.15
    for umbral in (0.15, 0.10, 0.07, 0.045, 0.03):
        candidatas, mask_u = _omr_bandas_por_gaps(col_smooth, bw, umbral)
        if len(candidatas) >= objetivo:
            bands, mask = candidatas, mask_u
            break
        if len(candidatas) > len(bands):
            bands, mask = candidatas, mask_u

    if len(bands) > max_bandas:
        # Encontrado con una foto real (smoke test fuera del repo): cuando el
        # bloque RESPUESTAS detectado queda un poco ancho, el texto vecino
        # ("USO EXCLUSIVO...") puede registrar suficiente tinta como para
        # aparecer como un candidato de banda MÁS ANCHO que las columnas
        # reales (que son angostas y uniformes entre sí). Quedarse con "los
        # más anchos" -- la lógica anterior -- descarta entonces una columna
        # real genuina (la más angosta del grupo) y conserva el bloque de
        # texto espurio en su lugar: la banda resultante queda desplazada una
        # posición completa (lo que el código llama "banda 3" termina siendo
        # físicamente la columna 4, y "banda 4" el texto vecino) -- un
        # desplazamiento de columna silencioso, sin relación con qué tan
        # ancha es cada una individualmente.
        #
        # Las columnas reales de esta plantilla son angostas y muy uniformes
        # entre sí (mismo layout impreso); un candidato espurio por contenido
        # vecino tiende a ser un outlier de ancho, no simplemente "el más
        # ancho de los reales". Por eso se conservan los `max_bandas`
        # candidatos cuyo ancho está MÁS CERCA de la mediana del grupo
        # completo -- esto descarta outliers en cualquier dirección (tanto
        # ruido angosto como bloques anchos espurios) en vez de asumir que
        # "ancho" siempre significa "real".
        anchos_cand = np.array([b[1] - b[0] for b in bands], dtype=np.float64)
        mediana_cand = float(np.median(anchos_cand))
        orden_cercania = np.argsort(np.abs(anchos_cand - mediana_cand))
        bands = [bands[i] for i in orden_cercania[:max_bandas]]
        bands = sorted(bands, key=lambda b: b[0])

    bandas_fabricadas = False
    if len(bands) < objetivo:
        if not permitir_reparto_geometrico:
            raise OMRError(
                f"Solo se distinguen {len(bands)} de las {objetivo} columnas esperadas y esta imagen "
                "es un recorte sin validar (no la hoja completa) -- no se fuerza un reparto geométrico "
                "porque podría fabricar una columna que no está realmente en la foto.")
        idx_contenido = np.where(col_smooth > 0.03)[0]
        x0c, x1c = (int(idx_contenido[0]), int(idx_contenido[-1]) + 1) if len(idx_contenido) else (0, bw)
        ancho = (x1c - x0c) / objetivo
        bands = [(int(x0c + i * ancho), int(x0c + (i + 1) * ancho)) for i in range(objetivo)]
        bandas_fabricadas = True

    return header_bottom, bands, bandas_fabricadas


def _omr_kmeans_1d(values, k, iters=50):
    values = np.asarray(values, dtype=np.float64)
    lo, hi = values.min(), values.max()
    centers = np.linspace(lo, hi, k) if hi > lo else np.full(k, lo)
    for _ in range(iters):
        d = np.abs(values[:, None] - centers[None, :])
        assign = d.argmin(axis=1)
        new_centers = centers.copy()
        for i in range(k):
            pts = values[assign == i]
            if len(pts) > 0:
                new_centers[i] = pts.mean()
        if np.allclose(new_centers, centers):
            break
        centers = new_centers
    return np.sort(centers)


def _omr_y_centers_uniforme(bh, n_filas):
    margen_y = bh * 0.02
    return np.linspace(margen_y, bh - margen_y, n_filas)


# ─────────────────────────── ROW LATTICE (v4.1) ───────────────────────────
# Reemplaza la asignación de los 20 centros de fila por kmeans 1D
# independiente (que no sabe que las filas impresas forman una estructura
# periódica y puede desplazar/duplicar/perder clusters, produciendo drift
# vertical acumulado dentro de una banda) por un modelo de grilla regular:
#
#     y(i) = y0 + i*dy      i = 0..n_filas-1
#
# y0 y dy se estiman de forma robusta a partir de los círculos Hough
# detectados en la banda (evidencia), no al revés -- los círculos NUNCA
# definen los 20 centros libremente, solo informan el origen y el período de
# una estructura que la plantilla impresa ya fija de antemano. Sobre esa
# grilla global se permite un microajuste local pequeño (±0.28*dy por fila,
# regularizado con una penalización de suavidad vía programación dinámica)
# para tolerar curvatura/perspectiva residual sin permitir que una fila
# individual se desplace lo suficiente como para ocupar la posición de otra.
def _omr_cluster_1d_ordenado(values_sorted, gap_min):
    """Agrupa un array 1D YA ordenado ascendente en clusters: un corte nuevo
    cuando el hueco entre valores consecutivos es >= gap_min. Devuelve una
    lista de (centro, n_puntos) en orden creciente de posición. Sirve tanto
    para estimar el período (dy) como para ubicar el origen (y0) de la
    grilla -- las burbujas A-E de una misma fila física caen todas dentro
    del mismo cluster porque su Y es casi idéntica."""
    if len(values_sorted) == 0:
        return []
    clusters = []
    start = 0
    for i in range(1, len(values_sorted)):
        if values_sorted[i] - values_sorted[i - 1] >= gap_min:
            grp = values_sorted[start:i]
            clusters.append((float(grp.mean()), int(len(grp))))
            start = i
    grp = values_sorted[start:]
    clusters.append((float(grp.mean()), int(len(grp))))
    return clusters


def _omr_estimar_dy_robusto(y_sorted, bh, n_filas, dy_esperado):
    """Estima el período vertical (dy) real entre filas a partir de las
    detecciones Hough de la banda, sin asumir que el primer/último círculo
    encontrado define nada. Método: agrupar en clusters de fila (ver arriba),
    tomar la mediana de las diferencias entre clusters consecutivos que caen
    en un rango plausible de "una fila" (0.5x a 1.6x el dy esperado por
    plantilla -- filtra huecos de 2+ filas por detecciones perdidas), y
    refinar una segunda vez con ese dy ya más preciso como umbral de
    clustering. Como señal adicional, contrasta contra dy implícito en el
    span primera-última fila detectada (item 14 del plan: evita que el error
    se acumule silenciosamente entre filas intermedias).

    Devuelve (dy, n_clusters_soporte, diagnostico). Si no hay evidencia
    suficiente para una estimación confiable, devuelve dy_esperado (prior de
    plantilla) sin fingir precisión que los datos no respaldan."""
    diag = {}
    gap_min0 = max(2.0, 0.45 * dy_esperado)
    clusters0 = _omr_cluster_1d_ordenado(y_sorted, gap_min0)
    centers0 = np.array([c for c, _n in clusters0])
    if len(centers0) < 3:
        return dy_esperado, len(centers0), {"motivo": "pocos_clusters_iniciales"}

    diffs0 = np.diff(centers0)
    validos0 = diffs0[(diffs0 >= 0.5 * dy_esperado) & (diffs0 <= 1.6 * dy_esperado)]
    dy1 = float(np.median(validos0)) if len(validos0) >= 3 else dy_esperado

    gap_min1 = max(2.0, 0.45 * dy1)
    clusters1 = _omr_cluster_1d_ordenado(y_sorted, gap_min1)
    centers1 = np.array([c for c, _n in clusters1])
    diffs1 = np.diff(centers1)
    validos1 = diffs1[(diffs1 >= 0.5 * dy1) & (diffs1 <= 1.6 * dy1)]
    dy2 = float(np.median(validos1)) if len(validos1) >= 3 else dy1

    dy_final = dy2
    if len(centers1) >= 2:
        span = centers1[-1] - centers1[0]
        k = max(1, round(span / dy2)) if dy2 > 0 else 1
        dy_span = span / k
        if dy_span > 0 and abs(dy_span - dy2) / dy2 < 0.25:
            dy_final = float(np.median([dy2, dy_span]))
        else:
            diag["desacuerdo_periodo"] = True  # ROW_PERIOD_ERROR: ver item 32

    return dy_final, len(centers1), diag


def _omr_estimar_y0_robusto(cluster_centers, dy, bh, n_filas, margen_frac=0.02):
    """Ubica el origen (y0, posición de la fila índice 0) usando la FASE
    circular de todos los clusters de fila detectados módulo dy (promedio
    circular -- robusto a outliers y no depende de "cuál círculo se detectó
    primero"), y ancla el múltiplo de dy correcto al margen superior
    esperado por la plantilla impresa (item 9: no basarse exclusivamente en
    el primer círculo encontrado ni en un porcentaje arbitrario aislado --
    acá el porcentaje de margen solo desempata ENTRE múltiplos de un período
    ya demostrado por evidencia real, no reemplaza la evidencia).

    El múltiplo elegido SIEMPRE debe dejar las n_filas físicas dentro de la
    banda (con un margen chico de tolerancia): un y0 que deja la fila 0 antes
    del borde superior, o la fila n_filas-1 después del inferior, no puede
    ser correcto -- no existe tinta impresa fuera de la banda. Sin este
    límite explícito, la fase circular por sí sola puede "empatar" entre dos
    múltiplos consecutivos de dy y quedarse con uno que saca la grilla de la
    banda por una fracción de fila (bug real, encontrado en smoke test con
    fotos reales: y0 <= -1px), lo que dispara los invariantes geométricos
    duros (centro_fuera_de_banda) y tira geometry_confidence a 0 innecesariamente."""
    if len(cluster_centers) == 0 or dy <= 0:
        return bh * margen_frac
    centers = np.asarray(cluster_centers, dtype=np.float64)
    phases = np.mod(centers, dy)
    angles = phases / dy * 2 * np.pi
    mean_angle = np.arctan2(np.sin(angles).mean(), np.cos(angles).mean())
    mean_phase = (mean_angle / (2 * np.pi)) * dy
    if mean_phase < 0:
        mean_phase += dy
    y0_prior = bh * margen_frac
    k = round((y0_prior - mean_phase) / dy)
    y0 = mean_phase + k * dy

    slack = dy * 0.5
    y0_min, y0_max = -slack, bh - (n_filas - 1) * dy + slack
    if y0_max < y0_min:  # banda mas baja que la grilla completa (dy sobreestimado): no hay margen valido
        return float(np.clip(y0, -slack, bh + slack))
    while y0 < y0_min:
        y0 += dy
    while y0 > y0_max:
        y0 -= dy
    return float(y0)


def _omr_row_support(y_pred, x_centers, obs_xy, tol_y, tol_x):
    """Soporte de una fila candidata: cuántas de las 5 columnas A-E tienen
    evidencia visual (un círculo Hough real) cerca de (x_columna, y_pred).
    Una fila real tiene A-E aproximadamente en el mismo Y (item 10) -- una
    detección aislada en una sola columna no puede por sí sola desplazar una
    fila completa. Devuelve un valor en [0, 5]."""
    if len(obs_xy) == 0:
        return 0.0
    near_y = obs_xy[np.abs(obs_xy[:, 1] - y_pred) <= tol_y]
    if len(near_y) == 0:
        return 0.0
    hits = 0
    for xc in x_centers:
        if np.any(np.abs(near_y[:, 0] - xc) <= tol_x):
            hits += 1
    return float(hits)


def omr_ajustar_lattice_vertical(obs_xy, x_centers, bh, n_filas, radio_banda, dy_prior=None, min_obs_total=3):
    """
    Núcleo del row lattice: reemplaza los 20 centros de fila independientes
    (antes kmeans 1D) por una grilla regular y(i) = y0 + i*dy + delta_i,
    donde y0/dy se estiman de forma robusta a partir de evidencia real y
    delta_i es un microajuste local acotado y suavizado (nunca "20 filas
    ajustadas por separado y después ordenadas" -- ver items 6-13 del plan).

    obs_xy: Nx2 array de (x, y) de círculos Hough detectados en esta banda
      (evidencia observacional -- pueden faltar filas enteras o sobrar
      outliers, la grilla NO depende de que estén todos).
    x_centers: los 5 centros de columna A-E ya resueltos (horizontal, fuera
      de alcance de este cambio).
    dy_prior: dy_global compartido entre las 4 bandas del mismo formulario
      (item 15) -- si el dy propio de esta banda se desvía demasiado o hay
      poca evidencia local, se prioriza el período compartido.

    Devuelve un dict: y_centers, y0, dy, row_support (uno por fila, 0..1),
    row_offsets (delta_i en píxeles), row_alignment_confidence (0..1),
    row_source ("ROBUST_LATTICE" | "LATTICE_INTERPOLATED" | "UNIFORM_FALLBACK"),
    diagnostico (dict con violaciones/señales para error analysis).
    """
    obs_xy = np.asarray(obs_xy, dtype=np.float64).reshape(-1, 2) if len(obs_xy) else np.empty((0, 2))
    x_centers = np.asarray(x_centers, dtype=np.float64)
    dy_esperado = bh / n_filas
    violaciones = []

    if len(obs_xy) < min_obs_total:
        y_centers = _omr_y_centers_uniforme(bh, n_filas)
        return {
            "y_centers": y_centers, "y0": float(y_centers[0]), "dy": dy_esperado,
            "row_support": np.zeros(n_filas), "row_offsets": np.zeros(n_filas),
            "row_alignment_confidence": 0.0, "row_source": "UNIFORM_FALLBACK",
            "diagnostico": {"violaciones": ["banda_sin_evidencia_circulos_lattice"]},
        }

    y_sorted = np.sort(obs_xy[:, 1])
    dy_local, n_clusters, dy_diag = _omr_estimar_dy_robusto(y_sorted, bh, n_filas, dy_esperado)
    if "desacuerdo_periodo" in dy_diag:
        violaciones.append("ROW_PERIOD_ERROR")

    dy = dy_local
    evidencia_debil = n_clusters < max(4, n_filas * 0.3)
    if dy_prior and dy_prior > 0:
        desvio = abs(dy_local - dy_prior) / dy_prior
        if evidencia_debil or desvio > 0.12:
            dy = dy_prior
            if desvio > 0.12 and not evidencia_debil:
                violaciones.append("dy_desvia_de_dy_global")

    gap_min = max(2.0, 0.45 * dy)
    clusters = _omr_cluster_1d_ordenado(y_sorted, gap_min)
    cluster_centers = np.array([c for c, _n in clusters])
    y0 = _omr_estimar_y0_robusto(cluster_centers, dy, bh, n_filas)

    tol_y = max(2.0, 0.35 * dy)
    tol_x = max(3.0, radio_banda * 1.6)

    # Chequeo anti ROW_INDEX_SHIFT (item 16): en una hoja real hay un círculo
    # IMPRESO en cada fila exista o no marca, así que la grilla es
    # aproximadamente periódica de punta a punta -- comparar el soporte TOTAL
    # de la ventana de 20 filas contra la ventana corrida ±1*dy es una señal
    # débil a propósito (correr toda la ventana un puesto solo cambia CUÁL
    # círculo real queda afuera en cada extremo, la enorme mayoría de las
    # filas interiores "coincide" en ambos casos por diseño). La señal que sí
    # distingue una ventana corrida de la correcta es LOCAL a los bordes: si
    # hay evidencia real de una fila más allá de un extremo, más fuerte que
    # la evidencia en ese extremo actual, la ventana entera está corrida.
    def _s(y):
        return _omr_row_support(y, x_centers, obs_xy, tol_y, tol_x)

    s_primera, s_antes_primera = _s(y0), _s(y0 - dy)
    s_ultima, s_despues_ultima = _s(y0 + (n_filas - 1) * dy), _s(y0 + n_filas * dy)
    ganancia_subir = s_antes_primera - s_ultima      # correr y0 -= dy: gana fila de arriba, pierde la última
    ganancia_bajar = s_despues_ultima - s_primera    # correr y0 += dy: gana fila de abajo, pierde la primera
    UMBRAL_SHIFT = 0.6  # en soporte 0..5 -- una fila real completa aporta ~5
    if ganancia_subir > UMBRAL_SHIFT and ganancia_subir >= ganancia_bajar:
        violaciones.append("ROW_INDEX_SHIFT_corregido")
        y0 -= dy
    elif ganancia_bajar > UMBRAL_SHIFT:
        violaciones.append("ROW_INDEX_SHIFT_corregido")
        y0 += dy
    elif abs(ganancia_subir) < UMBRAL_SHIFT and s_antes_primera > 0.6:
        # ni claramente mejor ni claramente peor corrida un puesto -- no se
        # puede afirmar con confianza cuál de las dos ventanas es la física
        # real (item 16: el orden y1<y2<... no alcanza como garantía).
        violaciones.append("ROW_INDEX_SHIFT_ambiguo")
    elif abs(ganancia_bajar) < UMBRAL_SHIFT and s_despues_ultima > 0.6:
        violaciones.append("ROW_INDEX_SHIFT_ambiguo")

    y_pred = y0 + np.arange(n_filas) * dy

    # Microajuste local acotado (±0.28*dy, ver item 12) + regularización por
    # suavidad vía programación dinámica (item 13): un salto grande en un
    # delta aislado solo se acepta si el soporte visual lo justifica lo
    # bastante como para pagar la penalización de discontinuidad.
    K = 13
    deltas_cand = np.linspace(-0.28 * dy, 0.28 * dy, K)
    soporte_grid = np.empty((n_filas, K))
    for i in range(n_filas):
        for k, d in enumerate(deltas_cand):
            soporte_grid[i, k] = _omr_row_support(y_pred[i] + d, x_centers, obs_xy, tol_y, tol_x)
    # Penalización mínima hacia delta=0 (además de la suavidad entre filas
    # vecinas de más abajo): sin esto, cuando la tolerancia de soporte es más
    # ancha que el rango de microajuste (tolerancia real de una foto borrosa,
    # o el caso límite sin ruido de estos tests sintéticos), TODOS los
    # candidatos de delta dan exactamente el mismo soporte para una fila --
    # un empate real que np.argmin rompe siempre hacia el primer índice
    # (el extremo -0.28*dy), produciendo un sesgo sistemático de toda la
    # banda hacia un mismo lado en vez de quedarse en delta=0 por falta de
    # evidencia que justifique moverse. La magnitud es deliberadamente chica
    # (a lo sumo ~0.04 en unidades de soporte 0..5) para que solo desempate,
    # nunca para que pese más que una diferencia real de soporte (perder o
    # ganar una sola columna ya vale 1.0).
    costo = -soporte_grid + 0.5 * (deltas_cand[None, :] / dy) ** 2
    lam = 3.0
    dp = np.empty((n_filas, K))
    back = np.zeros((n_filas, K), dtype=int)
    dp[0] = costo[0]
    penal_base = (deltas_cand[:, None] - deltas_cand[None, :]) ** 2 / (dy ** 2)
    for i in range(1, n_filas):
        total = dp[i - 1][None, :] + lam * penal_base
        back[i] = np.argmin(total, axis=1)
        dp[i] = costo[i] + total[np.arange(K), back[i]]
    idxs = np.zeros(n_filas, dtype=int)
    idxs[-1] = int(np.argmin(dp[-1]))
    for i in range(n_filas - 1, 0, -1):
        idxs[i - 1] = back[i, idxs[i]]
    row_offsets = deltas_cand[idxs]
    row_support = soporte_grid[np.arange(n_filas), idxs] / 5.0

    y_centers = y_pred + row_offsets
    # Límite físico duro: ninguna fila puede muestrearse fuera de la banda
    # (no existe tinta ahí). El anclaje de y0 ya deja margen para el
    # microajuste (±0.28*dy), pero un delta grande en la fila extrema todavía
    # puede empujarla un poco más allá del borde -- un clip final, no el
    # mecanismo primario, asegura que _omr_validar_invariantes_banda nunca
    # dispare "centro_fuera_de_banda" por este motivo.
    y_centers = np.clip(y_centers, 0.0, bh)

    # Dos niveles de soporte, no uno solo: una foto real de celular rara vez
    # deja que Hough encuentre las 5 burbujas A-E de la mayoría de las filas
    # (fotos reales de este smoke set rondan 10-20 círculos detectados de los
    # ~100 esperados por banda, aunque la geometría de fondo sea correcta) --
    # exigir "3 de 5 columnas" en al menos la mitad de las filas como piso de
    # confianza (como se probó primero) deja la confianza en 0 en CASI TODA
    # foto real, incluso cuando el ajuste es geométricamente consistente,
    # perdiendo por completo la utilidad del motor en fotos difíciles (la
    # app pasaba de "proponer con confianza media para revisar" a "no
    # proponer nada"). row_coverage_debil (al menos 1 columna real cerca de
    # la fila predicha) es la señal principal de densidad -- confirma que
    # ahí existe tinta real de esa fila, no que las 5 alternativas se vean.
    # row_coverage_fuerte (3+ columnas) sigue existiendo para distinguir el
    # nivel de evidencia en row_source, no como piso de confianza.
    row_coverage_debil = float(np.mean(row_support > 0))
    row_coverage_fuerte = float(np.mean(row_support >= 0.6))
    mean_abs_delta_norm = float(np.mean(np.abs(row_offsets)) / dy)
    if dy_prior and dy_prior > 0:
        dy_consistency = max(0.0, 1.0 - (abs(dy - dy_prior) / dy_prior) / 0.20)
    else:
        dy_consistency = 1.0 if not evidencia_debil else 0.5

    # row_alignment_confidence se combina MULTIPLICANDO con otros tres
    # factores ya existentes de geometry_confidence (ancho entre bandas,
    # invariantes duros, fuente de columnas) -- point recibido de un smoke
    # test con fotos reales: si este factor se arma "de a poco" (sumando
    # evidencia positiva hasta llegar a 1.0, como una primera version de
    # este calculo), CUALQUIER banda con evidencia sparse (la norma en una
    # foto real de celular, no la excepcion) queda con row_alignment_confidence
    # ~0.5-0.6 aun sin ninguna señal de mala alineación, y compuesto con los
    # otros tres factores (que ya rondan ~0.6-0.7 en fotos dificiles) el
    # producto final caía sistemáticamente por debajo del piso de
    # GEOMETRY_ERROR -- pasando de "revisar con una propuesta media" a "sin
    # propuesta alguna" en la enorme mayoría de fotos reales, no solo en las
    # que de verdad tienen un problema de alineación. Por eso el diseño es
    # "inocente hasta que se demuestre lo contrario": cada factor arranca
    # cerca de 1.0 y solo baja ante evidencia REAL de un problema (poca
    # densidad extrema, microajustes grandes, dy que no calza con las bandas
    # hermanas, o un shift de fila ambiguo) -- no por el mero hecho de que la
    # evidencia sea escasa pero consistente.
    f_densidad = 0.5 + 0.5 * min(1.0, row_coverage_debil / 0.5)
    f_suavidad = max(0.5, 1.0 - (mean_abs_delta_norm / 0.30) * 0.5)
    f_fuerte = 0.9 + 0.1 * row_coverage_fuerte
    row_alignment_confidence = f_densidad * f_suavidad * dy_consistency * f_fuerte
    if "ROW_INDEX_SHIFT_ambiguo" in violaciones:
        row_alignment_confidence *= 0.5
    if row_coverage_debil < 0.15:
        # Fail-closed (item 17), pero calibrado a evidencia real: si CASI
        # NINGUNA fila tiene siquiera una burbuja real cerca (banda
        # efectivamente vacía de evidencia, no solo escasa), no hay base
        # para afirmar que la grilla está alineada aunque período/origen
        # "parezcan" razonables.
        row_alignment_confidence = 0.0
    row_alignment_confidence = float(np.clip(row_alignment_confidence, 0.0, 1.0))
    row_coverage = row_coverage_debil  # nombre expuesto en diagnostico/tests

    # row_source describe qué tan demostrado quedó el resultado FINAL --
    # ROBUST_LATTICE exige evidencia densa de verdad (3+ columnas en la
    # mayoría de las filas), LATTICE_INTERPOLATED cubre el caso típico de
    # foto real (evidencia dispersa pero real y geométricamente consistente).
    if row_coverage_fuerte >= 0.7 and n_clusters >= max(6, n_filas * 0.4) and not evidencia_debil:
        row_source = "ROBUST_LATTICE"
    elif row_coverage_debil >= 0.15 or n_clusters >= 3:
        row_source = "LATTICE_INTERPOLATED"
    else:
        row_source = "UNIFORM_FALLBACK"

    return {
        "y_centers": y_centers, "y0": float(y0), "dy": float(dy),
        "row_support": row_support, "row_offsets": row_offsets,
        "row_alignment_confidence": row_alignment_confidence, "row_source": row_source,
        "diagnostico": {"violaciones": violaciones, "n_clusters": n_clusters,
                         "row_coverage": row_coverage, "mean_abs_delta_norm": mean_abs_delta_norm},
    }
# ───────────────────────── fin ROW LATTICE (v4.1) ─────────────────────────


# ───────────────────── NUMBER LATTICE (v4.1) ──────────────────────────────
# Segundo sensor geométrico INDEPENDIENTE de las burbujas: la plantilla
# imprime el número de cada fila (1..20 por banda) a la izquierda de la
# columna A. Sin OCR -- no importa qué dígito es, solo que hay tinta impresa
# ahí, igual que _omr_encontrar_barra_encabezado ya usa presencia de tinta
# (no lectura de texto) para ubicar la barra "RESPUESTAS". La franja de
# números tiene su propia periodicidad y0/dy, estimada con las MISMAS
# funciones robustas que el lattice de burbujas (reutilizadas, no
# reimplementadas) -- si ambos sensores coinciden, sube la confianza de que
# la correspondencia pregunta<->fila física es correcta; si discrepan en
# 0.5 filas o más, es evidencia fuerte de un ROW_INDEX_SHIFT que ninguno de
# los dos sensores por separado garantizaba poder detectar.
def _omr_franja_numeros(band_gray, x0_banda, xA, radio):
    """Sub-imagen de la franja vertical entre el borde izquierdo de la banda
    y la columna A (donde va impreso el número de fila), con un margen para
    no invadir la burbuja A misma."""
    x_fin = int(round(xA - radio * 1.5))
    x_ini = int(round(x0_banda))
    if x_fin - x_ini < 3:
        return None
    return band_gray[:, max(0, x_ini):max(x_ini + 3, x_fin)]


def _omr_detectar_number_y_obs(franja_gray):
    """Proyección vertical de tinta dentro de la franja de números: cada
    grupo contiguo de filas con oscuridad por encima del umbral es un
    candidato a "aquí hay un número impreso" -- mismo principio que
    _omr_encontrar_barra_encabezado (presencia de tinta, nunca lectura de
    texto), aplicado a muchos grupos chicos en vez de una sola barra."""
    if franja_gray is None or franja_gray.size == 0:
        return np.empty((0,))
    row_mean = franja_gray.mean(axis=1).astype(np.float64)
    umbral = row_mean.mean() - row_mean.std() * 0.4
    dark = row_mean < umbral
    obs = []
    in_run, start = False, 0
    h = len(dark)
    for y in range(h):
        if dark[y] and not in_run:
            in_run, start = True, y
        elif not dark[y] and in_run:
            in_run = False
            # Un run que TOCA el borde superior/inferior de la franja no
            # puede confirmarse como un número real -- puede ser el borde de
            # la banda o de la tabla filtrándose en el recorte (encontrado
            # en smoke test: un run "0..2" pegado al borde superior producía
            # una fase de origen falsa a pesar de que el resto de la franja
            # tenía evidencia limpia y regular). Un número real impreso deja
            # margen de papel en blanco antes/después, igual que la barra de
            # encabezado real (_omr_encontrar_barra_encabezado).
            if start > 0 and y - 1 < h - 1:
                obs.append((start + y - 1) / 2.0)
    return np.array(obs, dtype=np.float64)


def omr_ajustar_number_lattice(band_gray, x0_banda, xA, radio, bh, n_filas, dy_prior=None, min_obs=4):
    """Ajusta un lattice y0/dy independiente sobre la franja de números de
    fila (reutilizando _omr_estimar_dy_robusto/_omr_estimar_y0_robusto, ya
    genéricas sobre un array 1D de observaciones Y). Devuelve None en
    'y_centers' cuando la franja no tiene evidencia suficiente -- eso NO es
    un fallo del lattice de burbujas, simplemente este segundo sensor no
    pudo aportar nada acá (item 26: number strip borroso no debe penalizar
    injustificadamente)."""
    franja = _omr_franja_numeros(band_gray, x0_banda, xA, radio)
    obs_y = _omr_detectar_number_y_obs(franja) if franja is not None else np.empty((0,))
    dy_esperado = bh / n_filas
    if len(obs_y) < min_obs:
        return {"y_centers": None, "dy": dy_esperado, "y0": None,
                "number_alignment_confidence": 0.0, "n_obs": int(len(obs_y))}

    y_sorted = np.sort(obs_y)
    dy_local, n_clusters, _diag = _omr_estimar_dy_robusto(y_sorted, bh, n_filas, dy_esperado)
    dy = dy_local
    evidencia_debil = n_clusters < max(4, n_filas * 0.3)
    if dy_prior and dy_prior > 0 and (evidencia_debil or abs(dy_local - dy_prior) / dy_prior > 0.12):
        dy = dy_prior

    gap_min = max(2.0, 0.45 * dy)
    clusters = _omr_cluster_1d_ordenado(y_sorted, gap_min)
    cluster_centers = np.array([c for c, _n in clusters])
    if len(cluster_centers) < min_obs:
        return {"y_centers": None, "dy": dy, "y0": None,
                "number_alignment_confidence": 0.0, "n_obs": int(len(obs_y))}
    y0 = _omr_estimar_y0_robusto(cluster_centers, dy, bh, n_filas)
    y_centers = np.clip(y0 + np.arange(n_filas) * dy, 0.0, bh)

    tol = max(2.0, 0.4 * dy)
    hits = sum(1 for yc in y_centers if np.any(np.abs(cluster_centers - yc) <= tol))
    number_alignment_confidence = hits / n_filas
    return {"y_centers": y_centers, "dy": float(dy), "y0": float(y0),
            "number_alignment_confidence": float(number_alignment_confidence), "n_obs": int(len(obs_y))}


def omr_crosscheck_bubble_number(bubble_y_centers, number_lat, dy):
    """Compara el lattice de burbujas contra el de números de fila (item 10
    del plan): row_crosscheck_error[i] = abs(bubble_y[i]-number_y[i])/dy.
    Solo se considera evidencia real si el number lattice tiene su propia
    confianza mínima -- un number strip ilegible (confidence baja) no debe
    poder tirar abajo una geometría de burbujas por lo demás sana."""
    if number_lat.get("y_centers") is None or number_lat.get("number_alignment_confidence", 0) < 0.3:
        return {"row_crosscheck_error": None, "shift_confirmado": False}
    errores = np.abs(np.asarray(bubble_y_centers) - number_lat["y_centers"]) / dy
    mean_err = float(np.mean(errores))
    return {"row_crosscheck_error": mean_err, "shift_confirmado": mean_err >= 0.5,
            "por_fila": errores}
# ─────────────────────── fin NUMBER LATTICE (v4.1) ────────────────────────


def _omr_validar_invariantes_banda(y_centers, x_centers, radio, bh, bw, n_filas):
    """
    Invariantes geométricos duros sobre la grilla YA ajustada de una banda --
    protección estructural (no estadística) contra la misma familia del bug
    "USO EXCLUSIVO" y otros desplazamientos/columnas mal separadas: en una
    grilla real de 5 columnas (A-E) x n_filas sobre burbujas impresas, estas
    condiciones no pueden fallar. Se validan aparte de la consistencia de
    ancho entre bandas (que ya cubre `omr_ajustar_grilla`) porque son señales
    independientes -- una banda puede tener un ancho parecido a sus hermanas
    y aun así tener una grilla interna inválida (columnas colapsadas, filas
    fuera de la imagen, spacing irregular).

    Devuelve (multiplicador, violaciones): multiplicador en [0, 1] para
    aplicar sobre geometry_confidence (0.0 = geometría no demostrable, así
    haya "tinta oscura" ahí) y la lista de nombres de invariantes que
    fallaron o quedaron al límite, para diagnóstico (nunca para crashear).
    """
    violaciones = []
    multiplicador = 1.0
    x_centers = np.asarray(x_centers, dtype=np.float64)
    y_centers = np.asarray(y_centers, dtype=np.float64)

    # xA < xB < xC < xD < xE y y1 < y2 < ... < y_n: por construcción kmeans +
    # sort ya lo garantiza, pero un colapso de clusters (círculos agrupados
    # mal) puede dejar centros casi idénticos sin violar el orden estricto,
    # así que igual se revisa aparte más abajo vía spacing.
    if len(x_centers) != 5 or np.any(np.diff(x_centers) <= 0):
        violaciones.append("orden_columnas_A_E")
        multiplicador = 0.0
    if len(y_centers) != n_filas or np.any(np.diff(y_centers) <= 0):
        violaciones.append("orden_filas")
        multiplicador = 0.0

    # Spacing horizontal/vertical estable: layout impreso uniforme -> un
    # spacing muy irregular indica clusters que no corresponden a burbujas
    # reales en esa posición (p.ej. un outlier arrastrado por ruido).
    if len(x_centers) == 5:
        dx = np.diff(x_centers)
        cv_x = float(dx.std() / max(dx.mean(), 1e-6))
        if cv_x > 0.6:
            violaciones.append("spacing_horizontal_irregular")
            multiplicador = 0.0
        elif cv_x > 0.3:
            violaciones.append("spacing_horizontal_alerta")
            multiplicador = min(multiplicador, 0.6)
    if len(y_centers) == n_filas:
        dy = np.diff(y_centers)
        cv_y = float(dy.std() / max(dy.mean(), 1e-6))
        if cv_y > 0.6:
            violaciones.append("spacing_vertical_irregular")
            multiplicador = 0.0
        elif cv_y > 0.3:
            violaciones.append("spacing_vertical_alerta")
            multiplicador = min(multiplicador, 0.6)

    # ROIs dentro de límites de la imagen. El CENTRO de cada burbuja tiene que
    # estar dentro de la banda sin excepción (si no, la grilla está mal
    # ubicada). El círculo de MUESTREO de la fila/columna extrema sí puede
    # sobrepasar el borde por un margen chico sin que sea un problema real --
    # por diseño esas burbujas quedan pegadas al borde de la banda, así que
    # un overshoot del orden del radio es esperable, no una señal de error.
    if len(x_centers) and len(y_centers):
        centro_fuera = (x_centers.min() < 0 or x_centers.max() > bw or
                         y_centers.min() < 0 or y_centers.max() > bh)
        if centro_fuera:
            violaciones.append("centro_fuera_de_banda")
            multiplicador = 0.0
        else:
            overshoot = max(0.0,
                             radio - x_centers.min(), (x_centers.max() + radio) - bw,
                             radio - y_centers.min(), (y_centers.max() + radio) - bh)
            tolerancia = max(radio * 1.2, 2.0)
            if overshoot > tolerancia * 3:
                violaciones.append("roi_fuera_de_banda")
                multiplicador = 0.0
            elif overshoot > tolerancia:
                violaciones.append("roi_cerca_del_borde")
                multiplicador = min(multiplicador, 0.7)

    # Ninguna burbuja debería solaparse de forma absurda con su vecina: el
    # radio de muestreo no puede superar la mitad del spacing mínimo real.
    if len(x_centers) == 5 and len(y_centers) == n_filas and radio > 0:
        min_spacing = min(np.diff(x_centers).min(), np.diff(y_centers).min())
        if radio * 1.6 > min_spacing:
            violaciones.append("radio_excede_spacing")
            multiplicador = min(multiplicador, 0.5)

    return multiplicador, violaciones


def omr_ajustar_grilla(body_gray, bands, n_filas=OMR_N_FILAS_POR_BLOQUE, bandas_fabricadas=False):
    """Ajusta, PARA CADA BANDA POR SEPARADO, sus 5 centros de columna (kmeans 1D,
    sin cambios) y sus n_filas centros de fila -- estos últimos vía un ROW LATTICE
    (v4.1, ver `omr_ajustar_lattice_vertical`): y(i) = y0 + i*dy + delta_i, con
    y0/dy estimados de forma robusta a partir de los círculos Hough de la banda
    (nunca 20 posiciones independientes vía kmeans -- esa arquitectura no sabe que
    las filas impresas son una estructura periódica y puede acumular drift vertical
    dentro de una banda). Las 4 bandas comparten un `dy_global` (mediana entre las
    bandas con evidencia fuerte) que regulariza el dy propio de cada banda -- una
    banda borrosa se apoya en el período demostrado por sus hermanas en vez de caer
    directo a una grilla uniforme sin relación con la foto real.

    También devuelve, por banda, una geometry_confidence en [0, 1] con dos señales
    independientes:

    1. Consistencia de ancho entre bandas -- las 4 bandas del mismo formulario
       tienen el mismo layout impreso (20 filas x 5 letras), así que deberían
       medir aproximadamente lo mismo de ancho. Si el bloque RESPUESTAS
       detectado quedó demasiado ancho y una banda absorbió de más el margen en
       blanco o texto vecino ("USO EXCLUSIVO..."), esa banda va a ser notoriamente
       más ancha que sus hermanas -- una señal geométrica pura, independiente de
       la nitidez de la foto, que no necesita OCR para detectar el problema.
    2. Un piso mínimo de círculos Hough reales encontrados en la banda: si no hay
       PRÁCTICAMENTE NINGUNO (banda vacía de verdad, ej. una columna que la foto
       ni siquiera llegó a mostrar), no importa que el ancho sea consistente --
       ahí no hay nada que leer.

    Una foto borrosa/comprimida baja el conteo total de círculos que Hough logra
    encontrar en TODAS las bandas por igual (ninguna razón para desconfiar más de
    una banda que de otra por eso), así que el conteo de Hough se usa solo como
    piso de "banda completamente vacía", no como score graduado -- el score
    graduado real es la consistencia de ancho entre bandas hermanas.

    Además, sobre la grilla YA ajustada de cada banda se valida un tercer
    conjunto de señales independiente: los invariantes geométricos duros de
    `_omr_validar_invariantes_banda` (orden y spacing de columnas/filas, ROIs
    dentro de la banda, radio vs. spacing). Una banda puede tener un ancho
    parecido a sus hermanas y aun así tener una grilla interna inválida, así
    que las tres señales se combinan multiplicando (cualquiera puede llevar
    geometry_confidence a 0 por sí sola -- fail closed, no promedio).

    Un cuarto factor, `geometry_source`, deja trazabilidad explícita de qué
    mecanismo produjo la geometría de cada banda -- no todo lo que pasa las
    tres señales anteriores tiene el mismo respaldo real:
      - "HOUGH": los centros de fila/columna vienen de círculos Hough reales
        detectados en esa banda -- la evidencia más fuerte posible.
      - "UNIFORM_FALLBACK": Hough no encontró suficientes círculos EN ESA
        banda puntual y su grilla interna se repartió por proporciones (el
        límite x0,x1 de la banda en sí venía de huecos de tinta reales).
      - "GEOMETRIC_BAND_FALLBACK": ni siquiera el límite x0,x1 de la banda
        vino de huecos de tinta reales -- toda la imagen se repartió en
        columnas iguales porque `omr_detectar_header_y_bandas` no pudo
        distinguir los huecos reales (`bandas_fabricadas=True`). Es la
        evidencia más débil: ni el límite de la banda ni su grilla interna
        están demostrados, así que se penaliza más que UNIFORM_FALLBACK."""
    bh, bw = body_gray.shape
    blur = cv2.medianBlur(body_gray, 3)
    anchos = [x1 - x0 for x0, x1 in bands]
    min_band_w = min(anchos)
    ancho_mediano = float(np.median(anchos))
    r_guess = max(3, min_band_w // 12)
    circles = cv2.HoughCircles(
        blur, cv2.HOUGH_GRADIENT, dp=1, minDist=max(4, r_guess),
        param1=60, param2=14, minRadius=max(2, int(r_guess * 0.6)),
        maxRadius=int(r_guess * 1.8),
    )
    pts = circles[0] if circles is not None else np.empty((0, 3))
    esperados = 5 * n_filas
    # Penalización por fuente -- una geometría fabricada por fallback nunca
    # puede llegar al mismo geometry_confidence que una demostrada por
    # círculos reales, aunque pase las demás señales.
    MULT_POR_FUENTE = {"HOUGH": 1.0, "UNIFORM_FALLBACK": 0.7, "GEOMETRIC_BAND_FALLBACK": 0.45}

    # --- Pasada 1: por banda, columnas (x, sin cambios) + dy local propio ---
    banda_info = []
    for (x0, x1) in bands:
        sel = pts[(pts[:, 0] >= x0) & (pts[:, 0] < x1)]
        ancho_conf = min(x1 - x0, ancho_mediano) / max(x1 - x0, ancho_mediano, 1e-6)
        banda_vacia = len(sel) < max(3, esperados * 0.03)  # casi ningún círculo real -> no hay grilla ahí
        hough_ok = len(sel) >= max(5 * n_filas * 0.35, 10)
        if hough_ok:
            x_c = _omr_kmeans_1d(sel[:, 0], 5)
            r_banda = float(np.median(sel[:, 2]))
        else:
            bw_band = x1 - x0
            num_w = bw_band * 0.18
            x_c = x0 + num_w + (np.arange(5) + 0.5) * ((bw_band - num_w) / 5)
            r_banda = max(3.0, min_band_w * 0.08)
        dy_local, n_clusters = None, 0
        if len(sel) >= 3:
            dy_local, n_clusters, _diag = _omr_estimar_dy_robusto(np.sort(sel[:, 1]), bh, n_filas, bh / n_filas)
        banda_info.append({"x0": x0, "sel": sel, "ancho_conf": ancho_conf, "banda_vacia": banda_vacia,
                            "hough_ok": hough_ok, "x_c": x_c, "r_banda": r_banda,
                            "dy_local": dy_local, "n_clusters": n_clusters})

    # dy_global (item 15): las 4 bandas comparten el mismo layout impreso, así
    # que su período vertical debería ser muy parecido -- mediana entre las
    # bandas con evidencia fuerte (hough_ok), para que una banda borrosa se
    # apoye en sus hermanas en vez de caer directo a una grilla sin relación
    # con la foto real.
    dy_locales_fuertes = [b["dy_local"] for b in banda_info if b["hough_ok"] and b["dy_local"] is not None]
    dy_global = float(np.median(dy_locales_fuertes)) if dy_locales_fuertes else None

    # --- Pasada 2: fila (y) final vía row lattice robusto, con dy_global como prior ---
    y_centers_por_banda, band_x_centers, radios = [], [], []
    geometry_confidence, geometry_violaciones, geometry_source = [], [], []
    row_alignment_confidence_por_banda, row_source_por_banda = [], []
    for info in banda_info:
        sel, x_c, r_banda = info["sel"], info["x_c"], info["r_banda"]
        obs_xy = sel[:, :2] if len(sel) else np.empty((0, 2))
        lattice = omr_ajustar_lattice_vertical(obs_xy, x_c, bh, n_filas, r_banda, dy_prior=dy_global)
        y_c = lattice["y_centers"]

        y_centers_por_banda.append(y_c)
        band_x_centers.append(x_c)
        radios.append(r_banda)

        if bandas_fabricadas:
            fuente = "GEOMETRIC_BAND_FALLBACK"
        elif info["hough_ok"]:
            fuente = "HOUGH"
        else:
            fuente = "UNIFORM_FALLBACK"
        geometry_source.append(fuente)
        row_source_por_banda.append(lattice["row_source"])
        row_alignment_confidence_por_banda.append(round(lattice["row_alignment_confidence"], 3))

        mult_inv, violaciones = _omr_validar_invariantes_banda(y_c, x_c, r_banda, bh, bw, n_filas)
        if info["banda_vacia"]:
            violaciones = ["banda_sin_evidencia_circulos"] + violaciones
        violaciones = violaciones + lattice["diagnostico"].get("violaciones", [])
        mult_fuente = MULT_POR_FUENTE[fuente]

        # NUMBER LATTICE (segundo sensor, item 10-11 del plan): solo actúa
        # como VETO cuando confirma con evidencia propia razonable un
        # desacuerdo grande (>=0.5 filas) contra el lattice de burbujas --
        # nunca como bonus (evitar otro factor multiplicativo que erosione
        # la confianza de bandas por lo demás sanas, la misma lección del
        # ajuste anterior a row_alignment_confidence) y nunca si el propio
        # number lattice no tiene evidencia suficiente (number strip
        # borroso/ilegible no penaliza injustificadamente).
        mult_crosscheck = 1.0
        number_lat = omr_ajustar_number_lattice(body_gray, info["x0"], x_c[0], r_banda, bh, n_filas,
                                                  dy_prior=dy_global)
        crosscheck = omr_crosscheck_bubble_number(y_c, number_lat, lattice["dy"])
        if crosscheck["shift_confirmado"]:
            # Diagnóstico siempre visible (útil para smoke test / futura
            # habilitación), pero el veto sobre geometry_confidence respeta
            # el feature flag -- ver NUMBER_LATTICE_CROSSCHECK_HABILITADO.
            violaciones.append("BUBBLE_NUMBER_DISAGREEMENT" if NUMBER_LATTICE_CROSSCHECK_HABILITADO
                                else "BUBBLE_NUMBER_DISAGREEMENT_ignorado_flag_off")
            if NUMBER_LATTICE_CROSSCHECK_HABILITADO:
                mult_crosscheck = 0.25

        # row_alignment_confidence entra como un quinto factor multiplicativo
        # fail-closed (item 17): ninguna banda puede llegar a GEOMETRY_OK si
        # su alineación vertical no está demostrada, aunque las otras tres
        # señales (ancho, invariantes, fuente de columnas) estén perfectas.
        geo_conf = (0.0 if info["banda_vacia"] else
                    info["ancho_conf"] * mult_inv * mult_fuente * lattice["row_alignment_confidence"] * mult_crosscheck)
        geometry_confidence.append(geo_conf)
        geometry_violaciones.append(violaciones)

    radio = float(np.median(radios))
    return (np.array(y_centers_por_banda), np.array(band_x_centers), radio,
            geometry_confidence, geometry_violaciones, geometry_source,
            row_alignment_confidence_por_banda, row_source_por_banda)


def _omr_oscuridad_celda(gray, cx, cy, r, r_inner_frac=0.6):
    """
    Cuánto más oscuro está el INTERIOR de una burbuja que el fondo de papel en
    blanco. Antes se promediaba un parche cuadrado completo alrededor del
    centro, que mezcla el borde impreso del círculo, la letra impresa
    (A/B/C/D/E, que en esta plantilla queda cerca del borde) y el grafito
    real -- una burbuja VACÍA con su letra bien impresa podía dar casi la
    misma oscuridad promedio que una burbuja tenue pero realmente marcada.

    Ahora se recorta un parche del tamaño del radio completo (para tener
    contexto) pero se promedia solo dentro de una MÁSCARA CIRCULAR interna
    (r_inner = r_inner_frac * r, 0.6 por defecto -- dentro del rango 0.55-0.70
    esperable para que el borde y la letra impresa aporten mucho menos peso
    que el centro real de la burbuja, que es donde cae el grafito cuando se
    rellena)."""
    x0, x1 = max(0, int(cx - r)), int(cx + r + 1)
    y0, y1 = max(0, int(cy - r)), int(cy + r + 1)
    patch = gray[y0:y1, x0:x1]
    if patch.size == 0:
        return 0.0
    ph, pw = patch.shape
    yy, xx = np.ogrid[:ph, :pw]
    cy_local, cx_local = cy - y0, cx - x0
    r_inner = max(1.0, r * r_inner_frac)
    mask_inner = (xx - cx_local) ** 2 + (yy - cy_local) ** 2 <= r_inner ** 2
    if not mask_inner.any():
        return 255.0 - float(patch.mean())
    return 255.0 - float(patch[mask_inner].mean())


def omr_calcular_puntajes(body_gray, y_centers_por_banda, band_x_centers, radio):
    """Devuelve lista de dicts {A:.., B:.., ..} de oscuridad cruda por pregunta, en orden P1..Pn.
    El parche de cada celda usa el radio completo de la burbuja (más contexto que antes),
    pero la máscara interna dentro de _omr_oscuridad_celda es la que de verdad decide qué
    píxeles cuentan para el score."""
    sample_r = max(2.0, radio * 0.9)
    filas = []
    for bi, band_x in enumerate(band_x_centers):
        for cy in y_centers_por_banda[bi]:
            fila = {OMR_LETRAS[li]: _omr_oscuridad_celda(body_gray, cx, cy, sample_r)
                    for li, cx in enumerate(band_x)}
            filas.append(fila)
    return filas


def omr_clasificar_pregunta(scores, baseline, peak, umbrales=OMR_THRESHOLDS):
    letras = list(scores.keys())
    vals = np.array([scores[l] for l in letras])
    rng = max(peak - baseline, 1e-6)
    norm = (vals - baseline) / rng
    order = np.argsort(norm)[::-1]
    best_i, second_i = order[0], order[1]
    best_v, second_v = norm[best_i], norm[second_i]
    margin = best_v - second_v

    doble = best_v > umbrales["DOUBLE_MARK_THRESHOLD"] and second_v > umbrales["DOUBLE_MARK_THRESHOLD"]

    if best_v < umbrales["MIN_MARK_SCORE"]:
        return {"letra": None, "status": "sin_marca",
                "omr_confidence": round(max(0.0, min(1.0, float(best_v))), 3)}
    if doble:
        return {"letra": None, "status": "doble_marca",
                "omr_confidence": round(float(margin), 3)}
    if margin < umbrales["AMBIGUOUS_MARGIN"]:
        return {"letra": None, "status": "ambiguo",
                "omr_confidence": round(float(margin), 3)}
    alta = margin >= umbrales["HIGH_CONFIDENCE_MARGIN"]
    conf = max(0.0, min(1.0, 0.5 + margin))
    return {"letra": letras[best_i],
            "status": "alta_confianza" if alta else "confianza_media",
            "omr_confidence": round(float(conf), 3)}


def omr_recortar_pregunta(body_bgr, y_centers_por_banda, band_x_centers, radio, idx_local, n_bandas):
    """Recorta la fila completa (5 burbujas + contexto) de la pregunta con índice local
    idx_local (0-based, orden banda por banda), sobre la imagen SIN comprimir, para
    mostrársela a la persona en la UI de corrección manual."""
    n_filas = len(y_centers_por_banda[0])
    bi, ri = divmod(idx_local, n_filas)
    if bi >= len(band_x_centers):
        raise OMRError(f"idx_local={idx_local} cae fuera de las {len(band_x_centers)} bandas detectadas "
                        f"({n_filas} filas/banda) — la grilla no cubre esa pregunta.")
    band_x = band_x_centers[bi]
    cy = y_centers_por_banda[bi][ri]
    x0 = band_x[0] - radio * 3.2
    x1 = band_x[-1] + radio * 2.2
    y0 = cy - radio * 2.4
    y1 = cy + radio * 2.4
    h, w = body_bgr.shape[:2]
    x0, x1 = max(0, int(x0)), min(w, int(x1))
    y0, y1 = max(0, int(y0)), min(h, int(y1))
    return body_bgr[y0:y1, x0:x1]


OMR_COLOR_ALTA = (60, 180, 60)      # verde (BGR)
OMR_COLOR_MEDIA = (0, 200, 230)     # amarillo/naranjo
OMR_COLOR_AMBIGUA = (40, 40, 220)   # rojo
OMR_COLOR_BLANCO = (150, 150, 150)  # gris (sin marca, con confianza suficiente para no ser dudosa)
OMR_COLOR_IA = (200, 130, 0)        # azul (resuelta por la IA de apoyo tras quedar ambigua en el OMR)
OMR_COLOR_GEOMETRIA = (180, 0, 180)  # magenta (sin evidencia de grilla real -- posible zona equivocada)


def omr_anotar_diagnostico(body_bgr, y_centers_por_banda, band_x_centers, radio, resultados,
                            offset_pregunta=0, escala=3):
    """Genera la imagen de diagnóstico: un círculo alrededor de la burbuja elegida (o un
    punto junto a la fila si no se eligió ninguna) y una etiqueta "Pn:letra" coloreada
    según el estado final de cada pregunta. `escala` agranda la imagen para legibilidad."""
    base = cv2.resize(body_bgr, (body_bgr.shape[1] * escala, body_bgr.shape[0] * escala),
                       interpolation=cv2.INTER_NEAREST)
    dbg = base.copy()
    n_filas = len(y_centers_por_banda[0])
    for idx, r in enumerate(resultados):
        bi, ri = divmod(idx, n_filas)
        if bi >= len(band_x_centers):
            continue
        band_x = band_x_centers[bi]
        cy = y_centers_por_banda[bi][ri] * escala
        q = offset_pregunta + idx + 1
        status = r["status"]
        color = {
            "alta_confianza": OMR_COLOR_ALTA, "confianza_media": OMR_COLOR_MEDIA, "sin_marca": OMR_COLOR_BLANCO,
            "confiable": OMR_COLOR_ALTA, "revisar_media": OMR_COLOR_MEDIA,
            "revisar_dudoso": OMR_COLOR_AMBIGUA, "blanco": OMR_COLOR_BLANCO,
            "revisada_ia": OMR_COLOR_IA, "revisar_geometria": OMR_COLOR_GEOMETRIA, "geometry_error": OMR_COLOR_GEOMETRIA,
        }.get(status, OMR_COLOR_AMBIGUA)
        if r["letra"]:
            li = OMR_LETRAS.index(r["letra"])
            cx = band_x[li] * escala
            cv2.circle(dbg, (int(cx), int(cy)), int(radio * 1.3 * escala), color, max(1, escala // 2))
        else:
            cv2.circle(dbg, (int(band_x[0] * escala - radio * escala), int(cy)), 3 * escala, color, -1)
        conf_pct = round(r.get("omr_confidence", 0) * 100)
        label = f"{q}:{r['letra'] or '?'} {conf_pct}%"
        geo_state = r.get("geometry_state")
        if geo_state in ("GEOMETRY_WARNING", "GEOMETRY_ERROR"):
            # Solo se agrega geometry_confidence a la etiqueta cuando NO es
            # GEOMETRY_OK -- mostrarlo siempre en las 80 preguntas satura la
            # imagen; el caso que de verdad hay que poder verificar a simple
            # vista es justamente cuándo la geometría es dudosa o falló.
            label += f" g{round(r.get('geometry_confidence', 0) * 100)}%"
        cv2.putText(dbg, label, (int(band_x[0] * escala - radio * 3.4 * escala), int(cy) - int(radio * escala) - 2),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.32 * escala / 2, color, 1, cv2.LINE_AA)
    return dbg


OMR_ANCHO_REF_PAGINA = 1000   # ancho de referencia para localizar la tabla en la hoja completa
OMR_ANCHO_REF_TABLA = 480     # ancho de referencia para la tabla YA recortada (columnas/filas/círculos)
# Ambas referencias importan por separado: la tabla recortada es solo una
# fracción de la hoja completa (~4 columnas de burbujas), así que su propio
# ancho de referencia tiene que ser mucho más chico que el de la página
# entera -- usar el mismo valor para las dos etapas (bug real de una versión
# anterior de este fix) dejaba la segunda etapa corriendo a una escala varias
# veces más grande que aquella con la que se calibraron los parámetros de
# Hough, reproduciendo el mismo problema que se buscaba resolver.


def _omr_escalar_para_deteccion(img_bgr, ancho_referencia):
    """
    Devuelve (imagen_reescalada, escala) para las etapas de DETECCIÓN de
    geometría (contornos, huecos entre columnas, círculos de Hough). Nunca
    agranda -- solo achica si la imagen es más ancha que la referencia.

    Por qué existe: la app deliberadamente NUNCA reduce la resolución de la
    foto original antes de procesarla (para no perder detalle de las marcas a
    lápiz), pero una foto de celular moderna puede tener 3000-4000px de ancho,
    varias veces más que la imagen con la que se calibraron los parámetros de
    cv2.HoughCircles/medianBlur más abajo. Esos parámetros NO son invariantes
    a la escala: un blur de kernel fijo (3x3) es insignificante en una foto de
    alta resolución (deja pasar grano/textura del papel que Hough puede
    confundir con círculos), y los umbrales de Hough (param1/param2) fueron
    ajustados para círculos de cierto tamaño en píxeles. Probado de forma
    directa: la misma foto real reescalada a 4x resolución hizo caer la
    exactitud de 100% a ~21%, con solo 4 de 80 preguntas resueltas con
    confianza -- confirmando que el problema es de escala, no de la foto.

    La solución: la GEOMETRÍA (dónde están las columnas, filas y burbujas) se
    calcula siempre sobre una copia reescalada a un ancho de referencia fijo,
    y esas coordenadas se escalan de vuelta a la resolución original antes de
    MEDIR qué tan oscura está cada burbuja o de recortar una pregunta -- así
    la medición real sigue usando el detalle completo de la foto original.
    """
    h, w = img_bgr.shape[:2]
    if w <= ancho_referencia:
        return img_bgr, 1.0
    escala = ancho_referencia / w
    img_esc = cv2.resize(img_bgr, (int(w * escala), int(h * escala)), interpolation=cv2.INTER_AREA)
    return img_esc, escala


def omr_analizar_imagen(img_bgr, es_recorte, max_bandas=OMR_MAX_BANDAS, n_preguntas=None):
    """
    Punto de entrada principal del motor OMR: toma una imagen ya cargada (BGR,
    resolución original, SIN recomprimir) y devuelve la grilla ajustada +
    resultados crudos por pregunta local (orden banda por banda, 20 filas/banda).
    Ver _omr_escalar_para_deteccion() para por qué la geometría se detecta a
    una escala de referencia fija aunque la MEDICIÓN final use la foto entera
    a resolución original.

    es_recorte=False: se asume hoja completa -> se localiza el bloque
      RESPUESTAS y se endereza antes de leer la grilla.
    es_recorte=True: la imagen YA es (una porción de) la tabla RESPUESTAS ->
      se trabaja directo sobre toda la imagen.
    n_preguntas: si se indica, cuántas bandas (columnas de 20) hacen falta
      para cubrirlo -- permite el reparto geométrico proporcional cuando la
      foto no tiene contraste suficiente para distinguir los huecos reales.

    Lanza OMRError si no logra ubicar una grilla legible (la hoja queda sin
    resolver para completar a mano; nunca se reemplaza con una lectura de IA).
    """
    img = img_bgr
    if not es_recorte:
        img_det, escala_tabla = _omr_escalar_para_deteccion(img, OMR_ANCHO_REF_PAGINA)
        quad = omr_detectar_bloque_respuestas(img_det)
        if quad is None:
            raise OMRError("No se pudo localizar el bloque RESPUESTAS en la hoja completa.")
        if escala_tabla != 1.0:
            quad = quad / escala_tabla  # coordenadas de vuelta a la resolución original
        img = omr_enderezar_region(img, quad)  # perspective-warp sobre la imagen ORIGINAL, sin perder detalle

    bandas_esperadas = min(max_bandas, -(-n_preguntas // OMR_N_FILAS_POR_BLOQUE)) if n_preguntas else None

    # Referencia MUCHO más chica que la de arriba: acá "img" ya es solo la
    # tabla RESPUESTAS (4 columnas de burbujas), no la hoja completa.
    img_det, escala = _omr_escalar_para_deteccion(img, OMR_ANCHO_REF_TABLA)
    gray_det = cv2.cvtColor(img_det, cv2.COLOR_BGR2GRAY)
    header_bottom_det, bands_det, bandas_fabricadas = omr_detectar_header_y_bandas(
        gray_det, max_bandas=max_bandas, bandas_esperadas=bandas_esperadas,
        permitir_reparto_geometrico=not es_recorte)
    body_gray_det = gray_det[header_bottom_det:, :]
    (y_centers_det, band_x_centers_det, radio_det, geometry_confidence, geometry_violaciones, geometry_source,
     row_alignment_confidence_por_banda, row_source_por_banda) = (
        omr_ajustar_grilla(body_gray_det, bands_det, bandas_fabricadas=bandas_fabricadas))

    # Escalar toda la geometría encontrada de vuelta a la resolución original
    # antes de medir -- la medición de oscuridad y los recortes de preguntas
    # SIEMPRE usan la foto a resolución completa, nunca la copia reescalada.
    inv = 1.0 / escala
    header_bottom = int(round(header_bottom_det * inv))
    y_centers_por_banda = y_centers_det * inv
    band_x_centers = band_x_centers_det * inv
    radio = radio_det * inv

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    body_gray = gray[header_bottom:, :]
    body_bgr = img[header_bottom:, :]

    scores = omr_calcular_puntajes(body_gray, y_centers_por_banda, band_x_centers, radio)

    # baseline/peak POR BANDA, no globales: tolera iluminación despareja entre
    # columnas (sombra de la mano, ángulo de la luz en la foto).
    n_filas = len(y_centers_por_banda[0])
    resultados = []
    for bi in range(len(band_x_centers)):
        banda_scores = scores[bi * n_filas:(bi + 1) * n_filas]
        banda_vals = np.array([v for s in banda_scores for v in s.values()])
        baseline = float(np.percentile(banda_vals, 15))
        peak = float(np.percentile(banda_vals, 97))
        geo_conf = geometry_confidence[bi]
        if geo_conf < OMR_THRESHOLDS["MIN_GEOMETRY_CONFIDENCE"]:
            geo_estado = "GEOMETRY_ERROR"
        elif geo_conf < OMR_THRESHOLDS["WARNING_GEOMETRY_CONFIDENCE"]:
            geo_estado = "GEOMETRY_WARNING"
        else:
            geo_estado = "GEOMETRY_OK"
        for s in banda_scores:
            r = omr_clasificar_pregunta(s, baseline, peak)
            r["geometry_confidence"] = round(geo_conf, 3)
            r["geometry_state"] = geo_estado
            r["geometry_violaciones"] = geometry_violaciones[bi]
            r["geometry_source"] = geometry_source[bi]
            r["row_alignment_confidence"] = row_alignment_confidence_por_banda[bi]
            r["row_source"] = row_source_por_banda[bi]
            if geo_estado == "GEOMETRY_ERROR":
                # Sin evidencia real (círculos Hough + invariantes de grilla) de
                # que esta banda sea una grilla de burbujas -- pudo caer sobre
                # margen en blanco o texto vecino. No importa qué tan "clara"
                # parezca la marca: medir oscuridad sobre la zona equivocada
                # puede dar una lectura con apariencia de alta confianza pero
                # completamente falsa. Se fuerza a revisión manual en vez de
                # devolver una letra -- fail closed, nunca se fabrica geometría.
                r["letra"] = None
                r["status"] = "geometry_error"
            resultados.append(r)

    return {
        "body_bgr": body_bgr,
        "y_centers": y_centers_por_banda,
        "band_x_centers": band_x_centers,
        "radio": radio,
        "n_bandas": len(band_x_centers),
        "geometry_confidence_por_banda": [round(g, 3) for g in geometry_confidence],
        "geometry_violaciones_por_banda": geometry_violaciones,
        "geometry_source_por_banda": geometry_source,
        "row_alignment_confidence_por_banda": row_alignment_confidence_por_banda,
        "row_source_por_banda": row_source_por_banda,
        "resultados": resultados,
    }

# ═══════════════════════ fin motor OMR ═══════════════════════════════════


VERSION_APP = "4.2.0"
FECHA_ACTUALIZACION = "2026-08-11"
DESARROLLADO_POR = "Matías Rifo V."
# v4.2.0 (LOGIN + REGISTRO DE ACTIVIDAD): la plataforma pasa a requerir
# inicio de sesión -- grupo cerrado de cuentas precargadas por el
# administrador (ver gestionar_usuarios.py), sin auto-registro público. Cada
# login y cada exportación de Excel quedan respaldados en una hoja de
# cálculo externa (registro_sheets.py) -- SOLO METADATOS (usuario, curso,
# cantidad de hojas, contadores de confiables/dudosas): nunca nombres, RUT
# ni respuestas de estudiantes, que siguen viviendo únicamente en el Excel
# que el propio docente descarga. Streamlit Community Cloud no garantiza
# persistir archivos locales entre reinicios del contenedor, así que el
# respaldo vive fuera de la app, en Google Sheets, vía una cuenta de
# servicio (credenciales en st.secrets, nunca en el repo). Motor OMR sin
# cambios (OMR_ENGINE_VERSION se mantiene).
# v4.1.0 (PLATAFORMA): el sitio deja de ser un script de una sola pantalla y
# pasa a ser un router multipágina (st.navigation/st.Page) con una página de
# inicio "Herramientas Docentes" -- el Revisor de Hojas de Respuestas queda
# como la primera herramienta disponible, con espacio dejado para agregar
# más herramientas de IA para docentes (ver convención documentada junto al
# router, al final del archivo). Motor OMR sin cambios (OMR_ENGINE_VERSION
# se mantiene) -- este cambio es puramente de estructura/navegación de la
# app, no del motor de corrección.
# Motor OMR v4: geometry_confidence + geometry_state (OK/WARNING/ERROR) +
# geometry_source trazable (HOUGH/UNIFORM_FALLBACK/GEOMETRIC_BAND_FALLBACK)
# por banda, fail-closed en la localización de la tabla y en la selección de
# bandas (nunca se acepta una columna real más angosta a cambio de un
# candidato espurio más ancho), y arbitraje de IA para respuestas ambiguas
# apagado por defecto. Validado contra fixtures + smoke test con fotos
# reales adicionales (privado, fuera del repo) antes de fusionar a main.
#
# v4.1 (ROW LATTICE): reemplaza los 20 centros de fila por banda -- antes
# kmeans 1D independiente, sin noción de que las filas impresas son una
# estructura periódica, causa raíz del drift vertical acumulado dentro de
# una banda -- por una grilla regular y(i)=y0+i*dy+delta_i con y0/dy
# estimados de forma robusta (fase circular + anclaje al margen de
# plantilla, dy compartido entre bandas hermanas) y microajuste local
# acotado (±0.28*dy) y suavizado (programación dinámica). row_alignment_confidence
# entra como quinto factor fail-closed de geometry_confidence. Segundo
# sensor independiente (NUMBER LATTICE, sobre los números de fila impresos)
# implementado y con tests, pero su veto sobre geometry_confidence queda
# OFF por defecto (NUMBER_LATTICE_CROSSCHECK_HABILITADO=False) tras medir un
# falso positivo en la foto de calibración real -- no se habilita un gate
# adicional sin evidencia de que no introduce más riesgo del que evita.
#
# v4.1.1 (UBICACIÓN DE TABLA): causa raíz encontrada más arriba en el
# pipeline que el row lattice -- `omr_detectar_bloque_respuestas` podía
# fusionar la sección RESPUESTAS con secciones vecinas (identificación del
# alumno, N° de folleto) vía un cierre morfológico demasiado permisivo
# (iterations=2 → iterations=1), y aunque las secciones quedaran separadas,
# `_omr_candidato_tiene_header` solo confirmaba "hay una barra oscura
# arriba" sin distinguir la barra real "RESPUESTAS" de otras barras reales
# de la misma hoja (identificación, CÉDULA DE IDENTIDAD -- esta última una
# grilla de círculos real también, así que ni el conteo crudo de círculos
# alcanzaba). Corregido exigiendo densidad de filas reales inmediatamente
# bajo el header (`_omr_header_seguido_de_grilla`) y en el candidato
# completo (`_omr_candidato_evidencia_grilla`), con conteo de picos por
# máximos locales (`_omr_contar_picos_locales`) en vez de corridas sobre un
# piso fijo (que fusionaba filas reales adyacentes cuando el valle entre
# ellas no llegaba a cero). Validado contra 3 fotos reales privadas
# (fuera del repo): pasaron de leer casillas de nombre/cédula como si
# fueran respuestas -- o fallar por completo -- a leer las 80 preguntas
# con geometry_confidence 0.77-1.00 en la gran mayoría de bandas.
OMR_ENGINE_VERSION = "4.1.1"

st.markdown("""
<style>
.corr-badge {
    display:inline-block; padding:3px 8px; border-radius:6px;
    font-size:12px; font-weight:600; margin:2px;
}
.badge-ok   { background:#dcfce7; color:#166534; }
.badge-err  { background:#fee2e2; color:#991b1b; }
.badge-dud  { background:#fef9c3; color:#92400e; }
.badge-null { background:#f3f4f6; color:#6b7280; }
div[data-testid="stDataEditor"] { font-size:13px; }
div[data-testid="stFileUploader"] > div { min-height: 120px; }
</style>
""", unsafe_allow_html=True)

# ─── Estado de sesión ────────────────────────────────────────────────
def df_pauta_vacio(n: int) -> pd.DataFrame:
    return pd.DataFrame({
        "N°": [f"P{i}" for i in range(1, n + 1)],
        "Respuesta": pd.array([None] * n, dtype="object"),
    })

def safe_key(s: str) -> str:
    """Convierte nombre de archivo en clave segura para widgets."""
    return re.sub(r'[^a-zA-Z0-9]', '_', s)

for k, v in {
    "n_preguntas": 80,
    "resultados": {},
    "correcciones": {},      # {arch: {str(num_p): letra}}
    "info_edits": {},        # {arch: {campo: valor}}
    "pauta": [],
    "fotos_pendientes": {},  # {id_unico: {nombre, bytes, mime, hash}} — subidas sin procesar aún
    "pauta_df": df_pauta_vacio(80),
    "modo_captura": "completa",  # "completa" | "solo_respuestas"
    "usar_omr": True,             # motor OMR es el método principal; la IA solo apoya en dudas e identificación
    "ia_arbitraje_habilitado": False,  # OFF por defecto: OMR es la única fuente de A-E salvo que se active
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

OPCIONES = ["A", "B", "C", "D", "E", "—"]


def prompt_dinamico(n: int, num_imagenes: int = 1, solo_respuestas: bool = False) -> str:
    num_columnas = -(-n // 20)  # división hacia arriba: esta plantilla usa columnas de 20 preguntas
    rangos = []
    for c in range(num_columnas):
        ini = c * 20 + 1
        fin = min((c + 1) * 20, n)
        rangos.append(f"columna {c+1} = preguntas {ini} a {fin}")
    descripcion_columnas = "; ".join(rangos)

    if solo_respuestas and num_imagenes >= 2:
        nota_imagenes = """IMPORTANTE — te adjunto 2 fotos que son acercamientos en zoom de la MISMA hoja, ya
recortada de antemano para mostrar ÚNICAMENTE el bloque RESPUESTAS (sin cabecera ni datos del estudiante):
1. Acercamiento de la MITAD IZQUIERDA del bloque RESPUESTAS.
2. Acercamiento de la MITAD DERECHA del bloque RESPUESTAS (con algo de superposición con la anterior).

No hay ninguna otra imagen de referencia: estas 2 fotos son todo lo que tienes de esta hoja. No busques ni
inventes nombre, cédula o folleto — no aparecen en estas fotos y tampoco se piden en la respuesta.

"""
    elif solo_respuestas:
        nota_imagenes = """IMPORTANTE — la foto adjunta ya viene recortada de antemano para mostrar ÚNICAMENTE
el bloque RESPUESTAS (sin cabecera ni datos del estudiante). No busques ni inventes nombre, cédula o folleto —
no aparecen en esta foto y tampoco se piden en la respuesta.

"""
    elif num_imagenes >= 3:
        nota_imagenes = """IMPORTANTE — te adjunto 3 fotos de la MISMA hoja de respuestas, en este orden:
1. La hoja completa (úsala para identificar al estudiante: apellidos, nombres, cédula, folleto).
2. Un acercamiento en zoom de la MITAD IZQUIERDA de esa misma hoja.
3. Un acercamiento en zoom de la MITAD DERECHA de esa misma hoja (con algo de superposición con la anterior).

Para leer el bloque RESPUESTAS, usa SIEMPRE los acercamientos (imágenes 2 y 3), NUNCA la foto completa — en
los acercamientos cada burbuja se ve mucho más grande y es mucho más fácil distinguir cuál está marcada. La
foto completa (imagen 1) es solo de referencia general y para los datos de identificación del estudiante.

"""
    else:
        nota_imagenes = ""

    if solo_respuestas:
        bloque_estructura = f"""Estás viendo SOLO el bloque "RESPUESTAS" de una hoja de respuestas de alternativas
(A/B/C/D/E) de estudiantes chilenos, recortado de la plantilla fija "PLANTILLA DE HOJA DE RESPUESTAS". Esa
plantilla organiza siempre las respuestas en columnas de 20 preguntas cada una, puestas una al lado de la otra
de izquierda a derecha: {descripcion_columnas}. Dentro de cada columna, cada fila tiene el número de pregunta
impreso a la izquierda seguido de 5 burbujas A-E."""
    else:
        bloque_estructura = f"""Eres un experto en leer hojas de respuestas de alternativas (A/B/C/D/E) de estudiantes chilenos.
Todas las hojas que vas a revisar usan siempre la misma plantilla fija "PLANTILLA DE HOJA DE RESPUESTAS", con
esta estructura exacta:

- Arriba a la izquierda, "IDENTIFICACIÓN DEL ESTUDIANTE": Apellido Paterno, Apellido Materno y Nombres,
  escritos A MANO, una letra por casillero (no son burbujas).
- Arriba a la derecha, "CÉDULA DE IDENTIDAD": los dígitos pueden venir escritos a mano en la fila superior de
  esa sección y/o marcados con burbujas (una columna de burbujas 0-9 por cada dígito). Si el texto a mano es
  legible, úsalo como fuente principal del RUT; si no, complétalo leyendo qué burbuja está marcada en cada
  columna de dígito. Ignora el recuadro "PASAPORTE" (normalmente vacío).
- Más abajo a la izquierda, casilleros separados de "N° DE FOLLETO", "SEDE", "LOCAL" y "SALA" — a menudo
  vienen vacíos; si no hay nada escrito ahí, deja "nro_folleto" como cadena vacía.
- El bloque grande "RESPUESTAS" está dividido en columnas de 20 preguntas cada una, puestas una al lado de la
  otra de izquierda a derecha: {descripcion_columnas}. Dentro de cada columna, cada fila tiene el número de
  pregunta impreso a la izquierda seguido de 5 burbujas A-E."""

    if solo_respuestas:
        schema = """{
  "respuestas": ["A","B",...],
  "dudosas": [3, 15]
}"""
        reglas_formato = f"""Reglas de formato:
- "respuestas": exactamente {n} elementos, en el mismo orden P1..P{n}. Usa null solo si la pregunta está
  realmente omitida (ninguna marca visible), no como comodín para lo que no estés seguro.
- "dudosas": números de pregunta (1 a {n}) con riesgo real de error, según el criterio simple de arriba.
- Solo el JSON, sin explicación ni comentarios adicionales."""
    else:
        schema = """{
  "apellido_paterno": "...",
  "apellido_materno": "...",
  "nombres": "...",
  "cedula": "...",
  "nro_folleto": "...",
  "respuestas": ["A","B",...],
  "dudosas": [3, 15]
}"""
        reglas_formato = f"""Reglas de formato:
- "respuestas": exactamente {n} elementos, en el mismo orden P1..P{n}. Usa null solo si la pregunta está
  realmente omitida (ninguna marca visible), no como comodín para lo que no estés seguro.
- "dudosas": números de pregunta (1 a {n}) con riesgo real de error, según el criterio simple de arriba.
- Campos de texto ilegibles o no visibles en la hoja: cadena vacía "".
- Solo el JSON, sin explicación ni comentarios adicionales."""

    return f"""{nota_imagenes}{bloque_estructura}

CÓMO ESTÁ MARCADA ESTA HOJA:
Cada pregunta tiene 5 círculos impresos que dicen A, B, C, D, E. El estudiante marca su respuesta rellenando o
rayando con lápiz mina el círculo elegido. El círculo marcado se ve MÁS OSCURO, GRIS o RELLENO comparado con
los otros 4 de la misma fila, que quedan vacíos/blancos con solo la letra impresa adentro.

REGLA DE ORO para leer cada fila: para CADA fila de 5 burbujas, compara las 5 ENTRE SÍ y elige la que se vea
más oscura, gris o rellena respecto a sus vecinas de esa misma fila — no evalúes cada burbuja aislada ni le
pidas que esté "perfectamente" rellena. Incluso en una imagen borrosa o de baja resolución, el círculo marcado
casi siempre tiene visiblemente más tono gris que los vacíos de su misma fila; esa diferencia relativa dentro
de la fila es más confiable que juzgar una sola burbuja por sí sola.

PASO 1 — Usa el número IMPRESO de cada fila, nunca el orden espacial en que la vas mirando. Por ejemplo, la
fila que dice "21" al lado de las burbujas es la pregunta 21 sin importar en qué columna esté ni si la miraste
antes o después que la fila "5" — no asumas que "la siguiente fila hacia abajo" continúa la numeración de la
columna anterior; cada columna empieza y termina en el rango indicado arriba.

PASO 2 — Registra cada respuesta en el índice del arreglo que corresponde exactamente a su número de pregunta
impreso (posición 1 = pregunta "1", posición 21 = pregunta "21", etc.), recorriendo columna por columna.

PASO 3 — Antes de responder, verifica que el arreglo "respuestas" tenga exactamente {n} elementos, uno por
cada número de pregunta impreso del 1 al {n} sin saltos ni desplazamientos. Luego vuelve a mirar por segunda
vez, aplicando otra vez la REGLA DE ORO, SOLO las preguntas donde no quedaste 100% seguro de cuál opción
marcó el estudiante, y confírmalas con calma.

CUÁNDO usar null y marcar como "dudosa" (mantén esta lista lo más corta posible): únicamente cuando, al
comparar las 5 burbujas de una fila entre sí, las 5 se ven igual de vacías (pregunta omitida) o dos se ven
igual de oscuras/rellenas entre sí (ambigüedad real, no se puede distinguir cuál de las dos es la marca). Si
una burbuja es claramente la más oscura de su fila aunque sea levemente, ESA es la respuesta — no es dudosa,
y no la dejes en null. Nunca elijas una letra al azar "porque alguna debe ser": si tras aplicar la REGLA DE
ORO dos veces sigues sin poder distinguir cuál de las 5 está marcada, usa null y agrégala a "dudosas" en vez
de adivinar — adivinar en silencio es peor que un dudoso, porque se ve como una respuesta correcta pero puede
ser falsa.

Responde ÚNICAMENTE con un JSON válido, sin texto adicional ni markdown, con esta forma exacta:

{schema}

{reglas_formato}"""


def prompt_identificacion() -> str:
    """Prompt acotado para el motor OMR: la IA NO lee burbujas de respuestas aquí
    (eso ya lo resolvió el OMR), solo transcribe los datos manuscritos/marcados
    de la cabecera de la hoja — más barato y sin riesgo de confundir tareas."""
    return """Estás viendo la cabecera de una "PLANTILLA DE HOJA DE RESPUESTAS" chilena (no el
bloque de respuestas). Contiene:
- "IDENTIFICACIÓN DEL ESTUDIANTE": Apellido Paterno, Apellido Materno y Nombres, escritos A
  MANO, una letra por casillero.
- "CÉDULA DE IDENTIDAD": dígitos que pueden venir escritos a mano y/o marcados con burbujas
  (una columna de burbujas 0-9 por cada dígito). Usa el texto a mano si es legible; si no,
  completa leyendo qué burbuja está marcada en cada columna. Ignora "PASAPORTE" si está vacío.
- Casilleros de "N° DE FOLLETO", "SEDE", "LOCAL" y "SALA" — a menudo vacíos.

Responde ÚNICAMENTE un JSON válido, sin texto adicional:
{"apellido_paterno":"...","apellido_materno":"...","nombres":"...","cedula":"...","nro_folleto":"..."}

Campos ilegibles o no visibles: cadena vacía "". Solo el JSON."""


def prompt_revision_dudosas(preguntas: list) -> str:
    """
    Prompt acotado para la revisión de apoyo: el motor OMR ya resolvió toda la
    hoja y solo pide una segunda mirada en las filas puntuales donde no logró
    determinar con confianza cuál burbuja está marcada (marca débil, dos
    alternativas parecidas, o una mancha que se corre hacia la burbuja
    vecina). NO es el prompt completo de la hoja -- deliberadamente acotado a
    "cuál burbuja está marcada en este recorte", nunca a resolver la pregunta.
    """
    lista = ", ".join(f"P{p}" for p in preguntas)
    return f"""Te adjunto {len(preguntas)} recortes de una hoja de respuestas de alternativas
(A/B/C/D/E). Cada recorte muestra una o dos filas consecutivas: el número de pregunta impreso
a la izquierda y sus 5 burbujas A-E. Te los muestro en este orden: {lista} (si un recorte trae
dos filas, la pregunta que te interesa es la que coincide con ese número).

Un sistema automático de reconocimiento óptico ya leyó el resto de la hoja y solo tiene dudas
en estas preguntas puntuales. Tu única tarea es decidir, para cada una, cuál de las 5 burbujas
tiene la marca de lápiz -- no intentes resolver ni evaluar la pregunta.

Compara las 5 burbujas de esa fila ENTRE SÍ y elige la que se vea más oscura/rellena respecto a
sus vecinas. Si de verdad las 5 se ven vacías, responde null. Si dos están igual de marcadas y
es imposible distinguir cuál es la real, responde null también -- no adivines.

Responde ÚNICAMENTE un JSON válido con esta forma exacta, un elemento por recorte en el mismo
orden en que te los mostré:
{{"resultados": [{{"pregunta": N, "letra": "A"}}, {{"pregunta": N, "letra": null}}, ...]}}"""


# ─── Funciones de datos ──────────────────────────────────────────────

def api_key_activa() -> str:
    try:
        return st.secrets["ANTHROPIC_API_KEY"]
    except Exception:
        return st.session_state.get("api_key_input", "")

def tiene_secret() -> bool:
    try:
        _ = st.secrets["ANTHROPIC_API_KEY"]
        return True
    except Exception:
        return False

TIPOS_MIME = {"image/jpeg":"image/jpeg","image/jpg":"image/jpeg",
              "image/png":"image/png","image/webp":"image/webp","image/heic":"image/jpeg"}

def _img_a_b64_jpeg(img: Image.Image, lado_max: int, calidad: int) -> str:
    img = img.convert("RGB")
    if max(img.size) > lado_max:
        escala = lado_max / max(img.size)
        img = img.resize((max(1,int(img.width*escala)), max(1,int(img.height*escala))), Image.Resampling.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=calidad, optimize=True)
    return base64.standard_b64encode(buf.getvalue()).decode()

def mejorar_contraste_burbujas(img: Image.Image) -> Image.Image:
    """
    Solo para los acercamientos del bloque RESPUESTAS: pasar a escala de grises quita
    ruido de color de la foto (sombras, tono del papel) y subir contraste + nitidez
    separa más el gris del rayado a lápiz del blanco de la burbuja vacía. No se aplica
    a la foto completa porque ahí conviene mantener el color para leer bien nombre/RUT.
    """
    img_gris = img.convert("L").convert("RGB")
    img_contraste = ImageEnhance.Contrast(img_gris).enhance(1.8)
    return ImageEnhance.Sharpness(img_contraste).enhance(2.0)

def abrir_imagen_corregida(datos_bytes: bytes):
    """
    Abre la imagen a su resolución ORIGINAL (sin redimensionar ni comprimir) y
    corrige la orientación EXIF: las fotos de celular suelen traer el tag EXIF
    "Orientation" en vez de venir realmente rotadas en los píxeles, y sin
    corregir esto la imagen puede llegar "acostada" 90°/180° aunque se vea
    derecha en el celular, desalineando cualquier recorte/detección posterior.
    Esto NO cubre fotos que ya vienen genuinamente rotadas en los píxeles (sin
    metadato EXIF, p.ej. tras pasar por WhatsApp o un editor que lo eliminó) —
    para esas, la corrección tiene que hacerla quien toma/recorta la foto.
    Devuelve None si la imagen no se puede decodificar.
    """
    try:
        img = Image.open(io.BytesIO(datos_bytes))
        img = ImageOps.exif_transpose(img)
        return img.convert("RGB")
    except Exception:
        return None


def preparar_imagenes(datos_bytes: bytes, mime: str, solo_respuestas: bool = False):
    """
    Claude redimensiona internamente cualquier imagen a un techo fijo de resolución
    antes de analizarla. Si las 80 preguntas van en una sola foto, ese techo se
    reparte entre las 80 y cada burbuja queda en unos pocos píxeles — insuficiente
    para distinguir con fiabilidad el rayado a lápiz. Por eso se generan 2-3 versiones
    de la misma foto: mitad izquierda / mitad derecha del bloque RESPUESTAS (siempre),
    más la foto completa para identificación del alumno (solo si el modo NO es
    "solo_respuestas") — cada mitad le da a su franja del bloque RESPUESTAS su propio
    techo de resolución completo en vez de compartirlo con el resto de la hoja.
    """
    img = abrir_imagen_corregida(datos_bytes)
    if img is None:
        return [(base64.standard_b64encode(datos_bytes).decode(), mime)]
    w, h = img.size
    margen = 0.10  # superposición horizontal para no cortar una columna justo por la mitad
    mitad = w // 2
    overlap = int(w * margen)

    if solo_respuestas:
        # La foto ya viene recortada por quien la sube para mostrar solo el bloque
        # RESPUESTAS: no hay cabecera que saltar (top=0, se aprovecha el 100% del alto)
        # ni necesidad de una "completa" (la identificación es 100% manual en este modo),
        # así que todo el presupuesto de resolución de Claude se dedica a las burbujas.
        izq = mejorar_contraste_burbujas(img.crop((0, 0, min(w, mitad + overlap), h)))
        der = mejorar_contraste_burbujas(img.crop((max(0, mitad - overlap), 0, w, h)))
        return [
            (_img_a_b64_jpeg(izq, 1568, 92), "image/jpeg"),
            (_img_a_b64_jpeg(der, 1568, 92), "image/jpeg"),
        ]

    completa = _img_a_b64_jpeg(img, 1568, 90)
    # Recorte vertical conservador: en esta plantilla la identificación del alumno
    # siempre ocupa bastante más del 10% superior de la hoja, así que descartar solo
    # ese 10% en los acercamientos no arriesga cortar filas de RESPUESTAS aunque la
    # foto venga encuadrada de forma distinta cada vez.
    top = int(h * 0.10)
    izq = mejorar_contraste_burbujas(img.crop((0, top, min(w, mitad + overlap), h)))
    der = mejorar_contraste_burbujas(img.crop((max(0, mitad - overlap), top, w, h)))
    return [
        (completa, "image/jpeg"),
        (_img_a_b64_jpeg(izq, 1568, 92), "image/jpeg"),
        (_img_a_b64_jpeg(der, 1568, 92), "image/jpeg"),
    ]

def evaluar_sospecha(res: dict) -> None:
    """
    Red de seguridad estadística: si el modelo se desalinea de fila/columna (p.ej. por
    una foto rotada o muy inclinada) tiende a leer una franja completa como si fuera
    siempre la misma burbuja, produciendo una letra que se repite muchísimo más de lo
    que un patrón real de respuestas de examen permitiría. Esto no reemplaza la
    revisión humana, pero evita que un resultado así de inverosímil pase inadvertido
    como si fuera una lectura normal.
    """
    respuestas = res.get("respuestas", [])
    no_nulas = [r.upper() for r in respuestas if r]
    n_resp = len(no_nulas)
    if n_resp == 0:
        res["sospechoso"] = True
        res["motivo_sospecha"] = "No se detectó ninguna respuesta marcada en toda la hoja."
        return
    letra_top, freq_top = Counter(no_nulas).most_common(1)[0]
    proporcion = freq_top / n_resp
    if n_resp >= 15 and proporcion >= 0.55:
        res["sospechoso"] = True
        res["motivo_sospecha"] = (
            f"{round(proporcion*100)}% de las respuestas detectadas son '{letra_top}' "
            f"({freq_top} de {n_resp}) — patrón inusual para un examen real, probable "
            f"desalineación de fila/columna. Revisa esta hoja con la foto original.")
    else:
        res["sospechoso"] = False
        res["motivo_sospecha"] = ""

def _llamar_claude(cliente, bloques_imagen: list, n: int, num_imagenes: int,
                    solo_respuestas: bool = False, texto_extra: str = "") -> dict:
    msg = cliente.messages.create(
        model="claude-sonnet-5", max_tokens=6144,
        messages=[{"role":"user","content": bloques_imagen + [
            {"type":"text","text":prompt_dinamico(n, num_imagenes, solo_respuestas) + texto_extra},
        ]}],
    )
    if msg.stop_reason == "max_tokens":
        raise ValueError("La respuesta de la IA se cortó por límite de tokens antes de terminar el JSON.")
    texto = next((b.text for b in msg.content if b.type == "text"), None)
    if texto is None:
        raise ValueError("La respuesta de la IA no incluyó texto (solo bloques de razonamiento u otro tipo).")
    texto = texto.strip()
    m = re.search(r'\{[\s\S]*\}', texto)
    if m:
        texto = m.group(0)
    res = json.loads(texto)
    resp = res.get("respuestas", [])
    res["respuestas"] = (resp + [None]*n)[:n]
    res["dudosas"] = sorted(set(res.get("dudosas", [])))
    return res

REFUERZO_REINTENTO = (
    "\n\nATENCIÓN: en un intento anterior de leer esta misma hoja, el patrón de respuestas "
    "resultó estadísticamente inverosímil para un examen real (una misma letra se repitió "
    "muchísimo más de lo esperable), lo que sugiere que te desalineaste de fila o de columna "
    "en algún punto. Vuelve a examinar la hoja desde cero, con máximo cuidado, aplicando la "
    "REGLA DE ORO fila por fila y verificando el número IMPRESO de cada fila antes de registrar "
    "su respuesta.")

def procesar_imagen(cliente, nombre: str, datos_bytes: bytes, mime: str, n: int,
                     solo_respuestas: bool = False) -> dict:
    imagenes = preparar_imagenes(datos_bytes, mime, solo_respuestas)
    bloques_imagen = [
        {"type":"image","source":{"type":"base64","media_type":mt,"data":data}}
        for data, mt in imagenes
    ]
    res = _llamar_claude(cliente, bloques_imagen, n, len(imagenes), solo_respuestas)
    evaluar_sospecha(res)
    intentos = 1
    # Un solo reintento automático cuando el primer resultado se ve estadísticamente
    # inverosímil: es barato (una llamada más) frente al costo de que el profesor
    # confíe en un resultado erróneo sin darse cuenta.
    if res["sospechoso"]:
        res2 = _llamar_claude(cliente, bloques_imagen, n, len(imagenes), solo_respuestas, REFUERZO_REINTENTO)
        evaluar_sospecha(res2)
        intentos = 2
        if not res2["sospechoso"]:
            res = res2
        elif res2.get("respuestas") != res.get("respuestas"):
            # Ambos intentos siguen sospechosos: nos quedamos con el segundo (ya vio la
            # advertencia) pero el flag queda encendido para que quede marcada a mano.
            res = res2
    res["archivo"] = nombre
    res["n_preguntas"] = n
    res["intentos"] = intentos
    res["solo_respuestas"] = solo_respuestas
    return res


# ─── Motor OMR: única fuente de las respuestas ────────────────────────
# En vez de pedirle a Claude que lea las 400 burbujas de la hoja, un motor de
# visión clásica (funciones omr_* más arriba) mide directamente cuánto más
# oscura está cada burbuja que el papel en blanco, comparando dentro de la fila.
# Claude Vision NUNCA lee ni confirma burbujas — solo transcribe nombre/RUT de
# la cabecera (función aparte, más abajo). Las preguntas que el motor no logra
# determinar con confianza quedan "dudosas" con un recorte para revisión
# manual. Si el motor no logra procesar la hoja (tabla no localizable, imagen
# ilegible), esa hoja queda entera sin resolver para completar a mano — nunca
# se rellena con una lectura de IA. El flujo 100%-IA original (procesar_imagen,
# arriba) sigue disponible como alternativa manual desde el sidebar.

def _bgr_a_jpeg_b64(img_bgr, calidad: int = 95, lado_max: int = None):
    if lado_max and max(img_bgr.shape[:2]) > lado_max:
        escala = lado_max / max(img_bgr.shape[:2])
        img_bgr = cv2.resize(img_bgr, (max(1, int(img_bgr.shape[1]*escala)), max(1, int(img_bgr.shape[0]*escala))),
                              interpolation=cv2.INTER_AREA)
    ok, buf = cv2.imencode(".jpg", img_bgr, [cv2.IMWRITE_JPEG_QUALITY, calidad])
    if not ok:
        raise OMRError("No se pudo codificar un recorte OMR como JPEG.")
    return base64.standard_b64encode(buf.tobytes()).decode()


def llamar_claude_identificacion(cliente, header_bgr) -> dict:
    """Segunda llamada, pequeña y acotada: solo transcribe la cabecera (nombre/RUT/folleto),
    nunca lee burbujas de respuestas — eso ya lo resolvió el motor OMR."""
    data = _bgr_a_jpeg_b64(header_bgr, calidad=90, lado_max=1024)
    msg = cliente.messages.create(
        model="claude-sonnet-5", max_tokens=512,
        messages=[{"role": "user", "content": [
            {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": data}},
            {"type": "text", "text": prompt_identificacion()},
        ]}],
    )
    texto = next((b.text for b in msg.content if b.type == "text"), None)
    if texto is None:
        return {}
    m = re.search(r'\{[\s\S]*\}', texto.strip())
    try:
        return json.loads(m.group(0) if m else texto)
    except Exception:
        return {}


def llamar_claude_revision_dudosas(cliente, crops: list) -> dict:
    """
    Apoyo de IA SOLO para las preguntas que el motor OMR deja genuinamente
    dudosas (ambiguo/doble marca) -- no para las que ya resolvió con
    confianza. crops: lista de (numero_pregunta, imagen_bgr). Devuelve
    {numero_pregunta: letra_o_None}. Si la IA no da una respuesta clara para
    alguna pregunta, o la llamada falla, esa pregunta queda como None (sigue
    dudosa para revisión manual) -- nunca se adivina.
    """
    if not crops:
        return {}
    preguntas = [q for q, _ in crops]
    bloques = [
        {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg",
                                      "data": _bgr_a_jpeg_b64(c, calidad=95)}}
        for _, c in crops
    ]
    msg = cliente.messages.create(
        model="claude-sonnet-5", max_tokens=1024,
        messages=[{"role": "user", "content": bloques + [
            {"type": "text", "text": prompt_revision_dudosas(preguntas)},
        ]}],
    )
    texto = next((b.text for b in msg.content if b.type == "text"), None)
    resultado = {q: None for q in preguntas}
    if texto is None:
        return resultado
    m = re.search(r'\{[\s\S]*\}', texto.strip())
    try:
        data = json.loads(m.group(0) if m else texto)
        for item in data.get("resultados", []):
            q = item.get("pregunta")
            letra = item.get("letra")
            if q in resultado:
                resultado[q] = letra.upper() if isinstance(letra, str) and letra.upper() in LETRAS_VALIDAS else None
    except Exception:
        pass
    return resultado


def analizar_hoja_omr(datos_bytes: bytes, solo_respuestas: bool, n: int) -> dict:
    """
    Corre el motor OMR puro (sin llamadas a la API) sobre la imagen a resolución
    original. Devuelve la salida cruda de omr_analizar_imagen truncada a n
    preguntas, más la imagen (BGR, sin comprimir) por si hace falta recortar
    preguntas puntuales para mostrárselas a la persona durante la corrección
    manual. Lanza OMRError si no logra ubicar una grilla legible para las n
    preguntas configuradas (la llamante debe tratar esa hoja como no leíble).
    """
    img_pil = abrir_imagen_corregida(datos_bytes)
    if img_pil is None:
        raise OMRError("La imagen no se pudo decodificar.")
    img_bgr = cv2.cvtColor(np.array(img_pil), cv2.COLOR_RGB2BGR)
    salida = omr_analizar_imagen(img_bgr, solo_respuestas, n_preguntas=n)
    resultados = salida["resultados"]
    n_disponibles = len(resultados)
    if n_disponibles < n:
        # La tabla detectada cubre menos columnas de las que hacen falta para
        # n preguntas (p.ej. la foto no llegó a mostrar la 4ª columna, o la
        # detección de bandas no logró separarlas todas). Rellenar con
        # resultados "fantasma" es peligroso: no tienen banda/fila real en la
        # grilla, así que cualquier recorte o anotación posterior sobre esas
        # posiciones revienta con un índice fuera de rango. Es más seguro
        # tratar esta hoja como no leíble (queda para completar a mano).
        raise OMRError(
            f"La tabla detectada solo cubre {n_disponibles} de las {n} preguntas configuradas "
            "(probablemente la foto no muestra todas las columnas de RESPUESTAS).")
    salida["resultados"] = resultados[:n]
    salida["img_bgr_original"] = img_bgr
    salida["quad_respuestas"] = None if solo_respuestas else omr_detectar_bloque_respuestas(img_bgr)
    return salida


def _crop_dudosa_b64_jpeg(body_bgr, y_centers, band_x_centers, radio, idx_local, n_bandas, escala: int = 5) -> bytes:
    """Recorte ampliado de una pregunta dudosa, para mostrárselo a la persona en
    la UI de corrección (no se envía a ninguna IA)."""
    crop = omr_recortar_pregunta(body_bgr, y_centers, band_x_centers, radio, idx_local, n_bandas)
    crop_grande = cv2.resize(crop, (crop.shape[1] * escala, crop.shape[0] * escala),
                              interpolation=cv2.INTER_NEAREST)
    ok, buf = cv2.imencode(".jpg", crop_grande, [cv2.IMWRITE_JPEG_QUALITY, 90])
    return buf.tobytes() if ok else None


def _fallback_no_leido(nombre: str, n: int, solo_respuestas: bool, motivo: str) -> dict:
    """Resultado cuando el motor OMR no pudo leer una hoja, sea por lo que sea
    (tabla no localizable, imagen ilegible, o cualquier fallo inesperado más
    adelante en el pipeline). A propósito NO se cae a IA para inventar
    respuestas — nunca se le pide a Claude que lea burbujas. Todas las
    preguntas quedan sin resolver, para completar a mano o reprocesar esta
    foto puntual en modo "Solo IA" desde el sidebar si se prefiere."""
    return {
        "respuestas": [None] * n,
        "dudosas": list(range(1, n + 1)),
        "archivo": nombre, "n_preguntas": n, "solo_respuestas": solo_respuestas,
        "apellido_paterno": "", "apellido_materno": "", "nombres": "", "cedula": "", "nro_folleto": "",
        "omr_meta": {"usado": False, "motivo_fallback": motivo},
        # Se marca "sospechoso" igual que evaluar_sospecha() para que esta hoja
        # aparezca en la misma alerta y en la columna "Alerta calidad" del
        # Excel — si no, una hoja sin leer solo se nota abriéndola una por una.
        "sospechoso": True,
        "motivo_sospecha": "El motor OMR no pudo leer esta hoja — todas las preguntas quedaron sin resolver.",
    }


def procesar_imagen_hibrido(cliente, nombre: str, datos_bytes: bytes, mime: str, n: int,
                             solo_respuestas: bool = False, ia_arbitraje_habilitado: bool = False) -> dict:
    """
    Las respuestas SIEMPRE se determinan primero con el motor OMR (visión
    clásica) -- OMR es la fuente PRIMARIA, punto. Claude Vision solo entra en
    dos casos acotados, ninguno de los cuales es "leer una burbuja":
      1. Transcribir nombre/RUT cuando la imagen es la hoja completa
         (identificación) -- esto SIEMPRE está disponible, no depende de
         ningún toggle.
      2. Arbitrar preguntas ambiguas/doble marca -- SOLO si
         `ia_arbitraje_habilitado=True` (default False: apagado). Aun con el
         toggle encendido, Claude nunca puede tocar una pregunta cuya banda
         quedó en GEOMETRY_ERROR (no sabemos dónde están sus burbujas) ni
         reemplazar una respuesta que el OMR ya dio como confiable -- ver
         `_construir_resultado_omr`.
    Lo que ni el OMR ni ese apoyo puntual de IA logran determinar queda
    marcado como "dudosa" con un recorte ampliado adjunto para que la persona
    lo revise y corrija a mano en **Revisar y corregir**.
    """
    try:
        salida = analizar_hoja_omr(datos_bytes, solo_respuestas, n)
    except Exception as e:
        return _fallback_no_leido(nombre, n, solo_respuestas, str(e))

    try:
        return _construir_resultado_omr(salida, cliente, nombre, n, solo_respuestas, ia_arbitraje_habilitado)
    except Exception as e:
        # Cualquier fallo inesperado DESPUÉS de tener la grilla (p.ej. al
        # generar el diagnóstico visual o un recorte) tampoco debe tumbar la
        # hoja con un error crudo: se degrada al mismo resultado "no leído".
        return _fallback_no_leido(nombre, n, solo_respuestas, str(e))


def _construir_resultado_omr(salida: dict, cliente, nombre: str, n: int, solo_respuestas: bool,
                              ia_arbitraje_habilitado: bool = False) -> dict:
    resultados = salida["resultados"]
    body_bgr = salida["body_bgr"]
    y_centers, band_x_centers, radio = salida["y_centers"], salida["band_x_centers"], salida["radio"]
    umbral_blanco = OMR_THRESHOLDS["BLANK_REVIEW_MARGIN"]

    respuestas, dudosas, metodo_por_pregunta, confianza_por_pregunta = [], [], [], []
    ambiguas_para_ia = []  # subconjunto de dudosas con status crudo ambiguo/doble_marca: candidatas a apoyo de IA
    for i, r in enumerate(resultados):
        q = i + 1
        letra, status = r["letra"], r["status"]
        if status == "alta_confianza" and r.get("geometry_state") == "GEOMETRY_WARNING":
            # mark_confidence alto, pero geometry_confidence solo "aceptable con
            # reservas" (spacing irregular, radio cerca del límite, etc.) -- por
            # el principio geometry AND mark, no basta con que la marca se vea
            # clara para dar la pregunta como automática; se degrada a revisión
            # rápida en vez de "confiable" ciego.
            metodo = "revisar_media"
            dudosas.append(q)
        elif status == "alta_confianza":
            metodo = "confiable"
        elif status == "confianza_media":
            # se usa el propio mejor candidato del motor, pero queda marcada
            # como dudosa para que la persona la confirme de un vistazo
            metodo = "revisar_media"
            dudosas.append(q)
        elif status == "sin_marca" and r["omr_confidence"] <= umbral_blanco:
            metodo = "blanco"  # sin tinta suficiente ni para dudar: se confía en que está en blanco
        elif status == "geometry_error":
            # la banda de esta pregunta no tiene evidencia real (círculos Hough)
            # de ser una grilla de burbujas -- posiblemente cayó sobre margen en
            # blanco o texto vecino. No se manda a arbitraje de IA: un recorte de
            # esa zona no muestra una fila de alternativas real, así que pedirle
            # a Claude que elija una letra ahí solo arriesgaría inventar una.
            metodo = "revisar_geometria"
            letra = None
            dudosas.append(q)
        else:
            # ambiguo, doble_marca, o "sin_marca" con tinta suficiente para
            # sospechar una marca muy débil -> el motor no adivina; se intenta
            # una segunda mirada acotada de la IA antes de dejarla para revisión manual
            metodo = "revisar_dudoso"
            letra = None
            dudosas.append(q)
            # Candidata a arbitraje de IA solo si: (a) el toggle está
            # encendido -- default False, OMR es la fuente primaria y por
            # defecto ninguna IA toca respuestas; (b) el estado crudo es
            # ambiguo/doble_marca (nunca una marca débil sobre "sin_marca");
            # y (c) geometry_state == GEOMETRY_OK -- si la geometría de esta
            # banda no es sólida, no sabemos con certeza qué recorte le
            # estaríamos mostrando a Claude, así que ni se le pregunta
            # (geometry_error ya se filtra aparte arriba, pero GEOMETRY_WARNING
            # también queda excluido acá a propósito).
            if (ia_arbitraje_habilitado and status in ("ambiguo", "doble_marca")
                    and r.get("geometry_state") == "GEOMETRY_OK"):
                ambiguas_para_ia.append(q)
        respuestas.append(letra)
        metodo_por_pregunta.append(metodo)
        confianza_por_pregunta.append(r.get("omr_confidence", 0.0))

    crops_dudosas = {}
    crops_bgr_dudosas = {}
    for q in dudosas:
        try:
            crop_bgr = omr_recortar_pregunta(body_bgr, y_centers, band_x_centers, radio, q - 1, salida["n_bandas"])
            crop_bytes = _crop_dudosa_b64_jpeg(body_bgr, y_centers, band_x_centers, radio, q - 1, salida["n_bandas"])
            if crop_bytes:
                crops_dudosas[str(q)] = crop_bytes
            if crop_bgr is not None:
                crops_bgr_dudosas[q] = crop_bgr
        except Exception:
            pass  # sin recorte disponible para esta dudosa puntual; igual queda en la lista para corregir a mano

    # Apoyo de IA OPCIONAL (default apagado, ver ia_arbitraje_habilitado):
    # ambiguas_para_ia ya viene filtrada arriba a solo ambiguo/doble_marca con
    # geometry_state==GEOMETRY_OK -- nunca a lo que ya tiene una respuesta
    # confiable o de confianza media, y nunca a una banda sin geometría
    # demostrada. Si la IA tampoco logra decidir, la pregunta se deja igual
    # para revisión manual.
    n_api_calls_answer_arbitration = 0
    if ambiguas_para_ia and cliente is not None:
        crops_para_ia = [(q, crops_bgr_dudosas[q]) for q in ambiguas_para_ia if q in crops_bgr_dudosas]
        if crops_para_ia:
            n_api_calls_answer_arbitration = 1  # una sola llamada batcheada para todas las ambiguas de esta hoja
            try:
                resueltas_ia = llamar_claude_revision_dudosas(cliente, crops_para_ia)
            except Exception:
                resueltas_ia = {}
            for q, letra_ia in resueltas_ia.items():
                if letra_ia:
                    idx = q - 1
                    respuestas[idx] = letra_ia
                    metodo_por_pregunta[idx] = "revisada_ia"
                    dudosas.remove(q)

    # Conteo real de cómo se resolvió cada pregunta -- separado de las llamadas
    # a la API (arriba) porque antes omr_metrics.py asumía que la única
    # llamada posible del híbrido era la de identificación, sin contar la de
    # arbitraje de dudosas, lo que invalidaba esa métrica de costo (bug real,
    # corregido acá exponiendo ambos conteos por separado).
    n_answers_omr = sum(1 for m in metodo_por_pregunta if m in ("confiable", "revisar_media", "blanco"))
    n_answers_ai = sum(1 for m in metodo_por_pregunta if m == "revisada_ia")
    # No hay corrección manual dentro de este pipeline offline (ni en el harness
    # de métricas): "manual" ocurre después, en la UI, cuando la persona corrige
    # una dudosa -- este conteo siempre es 0 acá a propósito, no un placeholder
    # que se olvidó de llenar.
    n_answers_manual = 0
    n_answers_unresolved = sum(1 for m in metodo_por_pregunta if m in ("revisar_dudoso", "revisar_geometria"))

    res = {
        "respuestas": respuestas,
        "dudosas": sorted(set(dudosas)),
        "archivo": nombre,
        "n_preguntas": n,
        "solo_respuestas": solo_respuestas,
        "apellido_paterno": "", "apellido_materno": "", "nombres": "", "cedula": "", "nro_folleto": "",
        "omr_crops_dudosas": crops_dudosas,
        "omr_meta": {
            "usado": True,
            "n_bandas": salida["n_bandas"],
            "metodo_por_pregunta": metodo_por_pregunta,
            "confianza_por_pregunta": confianza_por_pregunta,
            "n_confiable": sum(1 for m in metodo_por_pregunta if m in ("confiable", "blanco")),
            "n_dudosas": len(dudosas),
            "geometry_confidence_por_banda": salida.get("geometry_confidence_por_banda", []),
            "geometry_violaciones_por_banda": salida.get("geometry_violaciones_por_banda", []),
            "geometry_source_por_banda": salida.get("geometry_source_por_banda", []),
            "row_alignment_confidence_por_banda": salida.get("row_alignment_confidence_por_banda", []),
            "row_source_por_banda": salida.get("row_source_por_banda", []),
            "n_geometry_error": sum(1 for m in metodo_por_pregunta if m == "revisar_geometria"),
            "n_geometry_warning": sum(1 for r in resultados if r.get("geometry_state") == "GEOMETRY_WARNING"),
            "n_answers_omr": n_answers_omr,
            "n_answers_ai": n_answers_ai,
            "n_answers_manual": n_answers_manual,
            "n_answers_unresolved": n_answers_unresolved,
            "n_api_calls_answer_arbitration": n_api_calls_answer_arbitration,
            "ia_arbitraje_habilitado": ia_arbitraje_habilitado,
            "n_api_calls_identification": 0,  # se corrige más abajo si de verdad se intenta la llamada
        },
    }

    # Datos del alumno: solo si la imagen es la hoja completa (en modo
    # solo_respuestas no hay cabecera que leer, igual que en el flujo 100%-IA)
    # y hay un cliente real disponible -- nunca toca respuestas, es la única
    # llamada de identificación en todo este pipeline.
    if not solo_respuestas and cliente is not None:
        res["omr_meta"]["n_api_calls_identification"] = 1
        quad = salida.get("quad_respuestas")
        img_bgr = salida["img_bgr_original"]
        top_tabla = int(min(p[1] for p in quad)) if quad is not None else int(img_bgr.shape[0] * 0.5)
        header_bgr = img_bgr[0:max(1, top_tabla), :]
        try:
            datos_id = llamar_claude_identificacion(cliente, header_bgr)
            for campo in ("apellido_paterno", "apellido_materno", "nombres", "cedula", "nro_folleto"):
                res[campo] = datos_id.get(campo, "") or ""
        except Exception:
            pass  # el alumno queda sin identificar automáticamente; se completa a mano, igual que hoy

    # Diagnóstico visual: útil para verificar de un vistazo qué detectó el motor
    # OMR en toda la hoja, sin tener que revisar pregunta por pregunta.
    resultados_diag = []
    for i, r in enumerate(resultados):
        rd = dict(r)
        rd["letra"] = respuestas[i]
        rd["status"] = metodo_por_pregunta[i]
        resultados_diag.append(rd)
    diag_bgr = omr_anotar_diagnostico(body_bgr, y_centers, band_x_centers, radio, resultados_diag)
    ok, buf = cv2.imencode(".jpg", diag_bgr, [cv2.IMWRITE_JPEG_QUALITY, 85])
    res["omr_diagnostico_bytes"] = buf.tobytes() if ok else None

    # evaluar_sospecha queda como ADVERTENCIA visible para la persona (banner en
    # la UI) — a propósito ya NO dispara un reprocesamiento con IA que reemplace
    # las respuestas del motor OMR por las de Claude.
    evaluar_sospecha(res)
    return res


def datos_efectivos(arch: str) -> dict:
    """Datos originales del alumno + ediciones manuales."""
    base = dict(st.session_state.resultados[arch])
    base.update(st.session_state.info_edits.get(arch, {}))
    return base

def respuestas_efectivas(arch: str) -> list:
    base = st.session_state.resultados[arch]["respuestas"].copy()
    for idx_str, letra in st.session_state.correcciones.get(arch, {}).items():
        idx = int(idx_str) - 1
        if 0 <= idx < len(base):
            base[idx] = None if letra == "—" else letra
    return base

def calcular(respuestas, pauta):
    co = inc = om = 0
    err = []
    for i, (r, p) in enumerate(zip(respuestas, pauta), 1):
        if not p: continue
        if r is None: om += 1
        elif r.upper() == p.upper(): co += 1
        else: inc += 1; err.append(i)
    return co, inc, om, err

LETRAS_VALIDAS = {"A", "B", "C", "D", "E"}

def pauta_desde_df(df: pd.DataFrame) -> list:
    return [
        r.strip().upper() if r and r.strip().upper() in LETRAS_VALIDAS else None
        for r in df["Respuesta"].fillna("").tolist()
    ]

def guardar_info_edit(arch, campo, valor):
    if arch not in st.session_state.info_edits:
        st.session_state.info_edits[arch] = {}
    st.session_state.info_edits[arch][campo] = valor.strip()

def guardar_correccion(arch, num_p, nueva, resp_orig):
    if nueva != (resp_orig or "—"):
        if arch not in st.session_state.correcciones:
            st.session_state.correcciones[arch] = {}
        st.session_state.correcciones[arch][str(num_p)] = nueva
    else:
        st.session_state.correcciones.get(arch, {}).pop(str(num_p), None)


# ─── Excel ───────────────────────────────────────────────────────────

def generar_excel(pauta: list, curso: str) -> bytes:
    n = len(pauta)
    wb = Workbook()
    bd = Border(left=Side(style="thin"),right=Side(style="thin"),
                top=Side(style="thin"),bottom=Side(style="thin"))
    AZ=PatternFill("solid",fgColor="2563EB"); AZD=PatternFill("solid",fgColor="1E3A5F")
    VE=PatternFill("solid",fgColor="DCFCE7"); RO=PatternFill("solid",fgColor="FECACA")
    AM=PatternFill("solid",fgColor="FEF08A"); AZC=PatternFill("solid",fgColor="DBEAFE")
    GR=PatternFill("solid",fgColor="F3F4F6")
    total_p = sum(1 for p in pauta if p)

    # Hoja 1 — Resumen (sin celdas combinadas: cada fila es autocontenida para
    # poder copiar/pegar y unificar varios Excel exportados en un archivo maestro,
    # filtrando por Curso y por alumno)
    ws1 = wb.active; ws1.title = "Resumen"
    enc = ["Curso","N°","Ap. Paterno","Ap. Materno","Nombres","Cédula","Folleto",
           "Correctas","Incorrectas","Omitidas","Puntaje %",
           "Preguntas incorrectas","Dudosas corregidas","Alerta calidad"]
    COL_PUNTAJE = 11
    COL_ALERTA = 14
    COLS_TEXTO = {1,3,4,5,6,7,14}
    fe = 1
    for c,h in enumerate(enc,1):
        cell=ws1.cell(fe,c,h); cell.font=Font(bold=True,color="FFFFFF")
        cell.fill=AZ; cell.alignment=Alignment(horizontal="center",wrap_text=True)
        cell.border=bd

    for rn,(arch,_) in enumerate(st.session_state.resultados.items(),1):
        datos = datos_efectivos(arch)
        ref   = respuestas_efectivas(arch)
        co,inc,om,err = calcular(ref, pauta[:len(ref)])
        pct = round(co/total_p*100,1) if total_p else 0
        res_orig = st.session_state.resultados[arch]
        dud = res_orig.get("dudosas",[])
        corr= st.session_state.correcciones.get(arch,{})
        cn  = len([k for k in corr if int(k) in dud])
        alerta = res_orig.get("motivo_sospecha","") if res_orig.get("sospechoso") else "—"
        vals=[curso or "—", rn,
              datos.get("apellido_paterno",""),datos.get("apellido_materno",""),
              datos.get("nombres",""),datos.get("cedula",""),datos.get("nro_folleto",""),
              co,inc,om,pct,
              ", ".join(str(e) for e in err) if err else "—",
              f"{cn} de {len(dud)}" if dud else "—",
              alerta]
        for c,v in enumerate(vals,1):
            cell=ws1.cell(fe+rn,c,v); cell.border=bd
            cell.alignment=Alignment(horizontal="left" if c in COLS_TEXTO else "center",wrap_text=True)
            if c==COL_PUNTAJE: cell.fill=VE if pct>=70 else (AM if pct>=50 else RO)
            if c==COL_ALERTA and alerta!="—": cell.fill=RO; cell.font=Font(bold=True,color="991B1B")
    for i,w in enumerate([20,4,16,16,22,14,9,10,11,10,10,38,18,44],1):
        ws1.column_dimensions[get_column_letter(i)].width=w
    ws1.row_dimensions[fe].height=22
    ws1.freeze_panes = "A2"

    # Hoja 2 — Detalle respuestas
    ws2 = wb.create_sheet("Detalle respuestas")
    cab=["Alumno"]+[f"P{i}" for i in range(1,n+1)]
    for c,h in enumerate(cab,1):
        cell=ws2.cell(1,c,h); cell.font=Font(bold=True,color="FFFFFF")
        cell.fill=AZD; cell.alignment=Alignment(horizontal="center"); cell.border=bd
    ws2.cell(2,1,"PAUTA").font=Font(bold=True)
    for i,p in enumerate(pauta,2):
        cell=ws2.cell(2,i,p or "?"); cell.fill=AZC
        cell.alignment=Alignment(horizontal="center"); cell.border=bd

    for fila,(arch,_) in enumerate(st.session_state.resultados.items(),3):
        datos=datos_efectivos(arch); ref=respuestas_efectivas(arch)
        corr=st.session_state.correcciones.get(arch,{})
        dud_set=set(st.session_state.resultados[arch].get("dudosas",[]))
        nombre=f"{datos.get('apellido_paterno','')} {datos.get('nombres','')}"
        ws2.cell(fila,1,nombre).border=bd
        for i,r in enumerate(ref,1):
            cell=ws2.cell(fila,i+1,r or "—")
            cell.alignment=Alignment(horizontal="center"); cell.border=bd
            corr_esta=str(i) in corr and i in dud_set
            if r is None: cell.fill=GR
            elif i<=len(pauta) and pauta[i-1] and r.upper()==pauta[i-1].upper():
                cell.fill=VE
                if corr_esta: cell.font=Font(bold=True,color="166534")
            else:
                cell.fill=RO
                if corr_esta: cell.font=Font(bold=True,color="991B1B")
            if i in dud_set and not corr_esta:
                cell.fill=AM; cell.font=Font(bold=True,color="92400E")
    ws2.column_dimensions["A"].width=26
    for col in range(2,n+2): ws2.column_dimensions[get_column_letter(col)].width=4.2
    ws2.row_dimensions[1].height=24

    # Hoja 3 — Estadísticas
    ws3=wb.create_sheet("Estadísticas")
    ws3["A1"]=f"Estadísticas — {curso or 'Curso'}"; ws3["A1"].font=Font(bold=True,size=14)
    ws3["A2"]=f"Total de preguntas evaluadas: {total_p}"
    todos_pct=[]
    for arch in st.session_state.resultados:
        ref=respuestas_efectivas(arch)
        co,_,_,_=calcular(ref,pauta[:len(ref)])
        todos_pct.append(round(co/total_p*100,1) if total_p else 0)
    for f,(e,v) in enumerate([
        ("Total alumnos",len(todos_pct)),
        ("Promedio (%)",round(sum(todos_pct)/len(todos_pct),1) if todos_pct else 0),
        ("Máximo (%)",max(todos_pct) if todos_pct else 0),
        ("Mínimo (%)",min(todos_pct) if todos_pct else 0),
        ("Sobre 60%",sum(1 for p in todos_pct if p>=60)),
        ("Bajo 60%",sum(1 for p in todos_pct if p<60)),
    ],4):
        ws3.cell(f,1,e).font=Font(bold=True); ws3.cell(f,2,v)
    ws3.column_dimensions["A"].width=34; ws3.column_dimensions["B"].width=14

    # Hoja 4 — Pauta
    ws4=wb.create_sheet("Pauta utilizada")
    ws4["A1"]=f"Pauta — {n} preguntas"; ws4["A1"].font=Font(bold=True,size=12)
    cols_f=5
    for idx,p in enumerate(pauta):
        fila=(idx//cols_f)+3; cb=(idx%cols_f)*2+1
        ws4.cell(fila,cb,f"P{idx+1}").font=Font(bold=True,color="1E3A5F")
        cell=ws4.cell(fila,cb+1,p or "—")
        cell.alignment=Alignment(horizontal="center")
        if p: cell.fill=AZC
    for col in range(1,cols_f*2+1): ws4.column_dimensions[get_column_letter(col)].width=7

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

def generar_qr(url: str) -> bytes:
    qr=qrcode.QRCode(box_size=6,border=2); qr.add_data(url); qr.make(fit=True)
    img=qr.make_image(fill_color="black",back_color="white")
    buf=io.BytesIO(); img.save(buf,format="PNG"); buf.seek(0)
    return buf.read()


def render_revisor_pruebas():
    """Renderiza la herramienta completa "Revisor de Hojas de Respuestas"
    (sidebar de configuracion + las 4 pestanas: Pauta, Cargar, Revisar,
    Exportar). Antes corria a nivel de modulo; el comportamiento interno
    es identico, solo quedo envuelto en una funcion para poder registrarse
    como una pagina mas del router multipagina (ver el bloque de navegacion
    al final del archivo).
    """
    # ═══════════════════════════════════════════════════════════════════════
    # SIDEBAR
    # ═══════════════════════════════════════════════════════════════════════
    with st.sidebar:
        st.markdown("## ⚙️ Configuración")
        if tiene_secret():
            st.success("🔑 API Key desde secrets")
        else:
            st.session_state["api_key_input"] = st.text_input(
                "API Key de Anthropic", type="password", placeholder="sk-ant-...")
        curso = st.text_input("Nombre del curso", placeholder="Ej: 3°A — Historia 2026")
        st.markdown("---")
        st.markdown("**Número de preguntas**")
        n_prev = st.session_state["n_preguntas"]
        hay_procesadas = bool(st.session_state["resultados"])
        n_nuevo = st.number_input("Preguntas", min_value=1, max_value=TEMPLATE_PROFILE["n_max"],
                                   value=n_prev, step=1, label_visibility="collapsed",
                                   disabled=hay_procesadas)
        if hay_procesadas:
            st.caption("🔒 Bloqueado: ya hay hojas procesadas con este N°. "
                       "Usa **Limpiar todo** antes de cambiarlo (cambiarlo a medio camino "
                       "descuadra el puntaje de las hojas ya analizadas).")
        elif n_nuevo != n_prev:
            st.session_state["n_preguntas"] = n_nuevo
            st.session_state["pauta_df"] = df_pauta_vacio(n_nuevo)
            st.session_state["pauta"] = []
            st.rerun()
        st.caption(f"Configurado para **{st.session_state['n_preguntas']} preguntas**")
        st.markdown("---")
        st.markdown("**Modo de carga de fotos**")
        opciones_modo = {
            "completa": "📄 Hoja completa (recorte automático)",
            "solo_respuestas": "✂️ Solo bloque RESPUESTAS (ya recortado por ti)",
        }
        modo_prev = st.session_state["modo_captura"]
        modo_nuevo = st.radio(
            "Modo de carga", options=list(opciones_modo.keys()),
            format_func=lambda k: opciones_modo[k],
            index=list(opciones_modo.keys()).index(modo_prev),
            label_visibility="collapsed", disabled=hay_procesadas,
        )
        if hay_procesadas:
            st.caption("🔒 Bloqueado: ya hay hojas procesadas con este modo. Usa **Limpiar todo** antes de cambiarlo.")
        elif modo_nuevo != modo_prev:
            st.session_state["modo_captura"] = modo_nuevo
            st.rerun()
        if st.session_state["modo_captura"] == "solo_respuestas":
            st.caption("✂️ Nombre, RUT y folleto quedan en blanco — se completan a mano.")
        else:
            st.caption("📄 La app recorta e identifica al alumno automáticamente.")
        st.markdown("---")
        st.markdown("**Motor de lectura**")
        if not OMR_DISPONIBLE:
            st.caption("🤖 Solo IA (motor OMR no disponible en este entorno).")
        else:
            omr_prev = st.session_state["usar_omr"]
            omr_nuevo = st.toggle(
                "🔬 Motor OMR — única fuente de las respuestas",
                value=omr_prev, disabled=hay_procesadas,
                help="Mide directamente qué tan oscura está cada burbuja (sin llamar a la API). Claude "
                     "nunca lee burbujas, solo transcribe nombre/RUT. Lo que el motor no logra determinar "
                     "con confianza queda dudoso, con su recorte, para corregir a mano — nunca se adivina.",
            )
            if hay_procesadas:
                st.caption("🔒 Bloqueado: ya hay hojas procesadas con este motor. Usa **Limpiar todo** antes de cambiarlo.")
            elif omr_nuevo != omr_prev:
                st.session_state["usar_omr"] = omr_nuevo
                st.rerun()
            if not st.session_state["usar_omr"]:
                st.caption("🤖 Modo 100% IA: más lento, más caro y menos confiable — solo para comparar.")
            elif st.session_state["usar_omr"]:
                ia_arb_prev = st.session_state["ia_arbitraje_habilitado"]
                ia_arb_nuevo = st.checkbox(
                    "🤖 Usar IA para arbitrar respuestas ambiguas (experimental)",
                    value=ia_arb_prev, disabled=hay_procesadas,
                    help="APAGADO por defecto: el OMR es la única fuente de A-E. Si se activa, Claude solo "
                         "puede opinar sobre preguntas ambiguas/doble marca cuya geometría es sólida "
                         "(GEOMETRY_OK) -- nunca sobre una respuesta ya confiable, ni sobre una banda sin "
                         "evidencia geométrica real (GEOMETRY_ERROR).",
                )
                if hay_procesadas:
                    st.caption("🔒 Bloqueado: ya hay hojas procesadas. Usa **Limpiar todo** antes de cambiarlo.")
                elif ia_arb_nuevo != ia_arb_prev:
                    st.session_state["ia_arbitraje_habilitado"] = ia_arb_nuevo
                    st.rerun()
        st.markdown("---")
        if st.button("🗑️ Limpiar todo", use_container_width=True):
            nn = st.session_state["n_preguntas"]
            st.session_state.update({
                "resultados":{}, "correcciones":{}, "info_edits":{},
                "fotos_pendientes":{}, "pauta":[], "pauta_df":df_pauta_vacio(nn),
            })
            st.rerun()
        st.caption(f"Versión {VERSION_APP} · OMR {OMR_ENGINE_VERSION} · Actualizado {FECHA_ACTUALIZACION}")
        st.caption(f"Desarrollado por {DESARROLLADO_POR}")


    # ═══════════════════════════════════════════════════════════════════════
    # CUERPO
    # ═══════════════════════════════════════════════════════════════════════
    n = st.session_state["n_preguntas"]
    st.markdown("# 📝 Revisor de Hojas de Respuestas")

    tab_pauta, tab_cargar, tab_revisar, tab_exportar = st.tabs([
        "📋 Pauta de respuestas",
        "📤 Cargar fotos",
        "✏️ Revisar y corregir",
        "📊 Exportar Excel",
    ])

    # ══ PAUTA ════════════════════════════════════════════════════════════
    with tab_pauta:
        st.markdown(f"### Respuestas correctas — {n} preguntas")
        st.caption("💡 Usa **Importar rápido** para pegar todas de una vez; la tabla solo para corregir 1-2 celdas.")
        col_t, col_i = st.columns([2, 1], gap="large")

        with col_i:
            st.markdown("**Importar rápido**")
            texto_bulk = st.text_area("Importar", height=200, label_visibility="collapsed",
                placeholder=f"A\nB\nC\n...\no bien: A,B,C,D,E,...\n({n} respuestas)")
            if st.button("⬇️ Cargar al grid", use_container_width=True):
                letras=[l.upper() for l in re.findall(r'[AaBbCcDdEe]', texto_bulk)]
                if letras:
                    st.session_state["pauta_df"] = pd.DataFrame({
                        "N°":[f"P{i}" for i in range(1,n+1)],
                        "Respuesta": pd.array((letras+[None]*n)[:n], dtype="object"),
                    })
                    st.success(f"✓ {min(len(letras),n)} respuestas cargadas")
                    st.rerun()
                else:
                    st.error("No se encontraron letras A–E válidas.")
            st.markdown("---")
            pauta_actual = pauta_desde_df(st.session_state["pauta_df"])
            completadas  = sum(1 for p in pauta_actual if p)
            st.metric("Con respuesta", f"{completadas} / {n}")
            if completadas == n:
                st.success("✅ Pauta completa")
            elif completadas > 0:
                falt=[i+1 for i,p in enumerate(pauta_actual) if not p]
                st.warning(f"Faltan {n-completadas}: {', '.join(f'P{x}' for x in falt[:5])}{'...' if len(falt)>5 else ''}")

        with col_t:
            df_ed = st.data_editor(
                st.session_state["pauta_df"],
                column_config={
                    "N°": st.column_config.TextColumn("N°", disabled=True, width="small"),
                    "Respuesta": st.column_config.SelectboxColumn(
                        "Respuesta", options=["A","B","C","D","E"], width="small", required=False),
                },
                hide_index=True, height=min(38*n+40, 560),
                use_container_width=True, key="editor_pauta", num_rows="fixed",
            )
            st.session_state["pauta_df"] = df_ed
            st.session_state["pauta"]    = pauta_desde_df(df_ed)

        if any(st.session_state["pauta"]):
            with st.expander("👁️ Vista previa"):
                html="<div style='line-height:2.2;'>"
                for i,p in enumerate(st.session_state["pauta"],1):
                    bg="#dcfce7" if p else "#f3f4f6"; fg="#166534" if p else "#9ca3af"
                    html+=(f'<span style="display:inline-block;width:50px;margin:2px;'
                           f'background:{bg};color:{fg};border-radius:4px;'
                           f'padding:2px 4px;font-size:12px;font-weight:600;">'
                           f'P{i}:{p or "?"}</span>')
                html+="</div>"
                st.markdown(html, unsafe_allow_html=True)


    # ══ CARGAR FOTOS ═════════════════════════════════════════════════════
    with tab_cargar:
        try:
            app_url = st.secrets.get("APP_URL","")
        except Exception:
            app_url = ""

        modo_actual = st.session_state["modo_captura"]

        if app_url:
            cq, ci = st.columns([1,2], gap="large")
            with cq:
                st.markdown("### 📱 Desde el celular")
                st.image(generar_qr(app_url), width=180, caption="Escanea para abrir en tu celular")
            with ci:
                st.markdown("### Instrucciones")
                if modo_actual == "solo_respuestas":
                    st.markdown("""
    1. Escanea el QR o entra a la URL en el navegador del celular
    2. **Cargar fotos** → **Upload** → Cámara o Galería
    3. Encuadra **solo el bloque RESPUESTAS completo** (las 4 columnas, sin cabecera), foto
       derecha y con buena luz
    4. Sube varias antes de procesar; nombre y RUT se completan a mano después
    """)
                else:
                    st.markdown("""
    1. Escanea el QR o entra a la URL en el navegador del celular
    2. **Cargar fotos** → **Upload** → Cámara o Galería
    3. Acerca la cámara a la hoja completa, con **las 4 columnas de RESPUESTAS visibles**,
       foto derecha y con buena luz
    4. Sube varias antes de procesar
    """)
        else:
            st.info("Desde el celular: entra a la URL de esta app; al tocar Upload ofrece abrir la "
                    "cámara directamente. Agrega `APP_URL` a Secrets para ver un QR aquí.")

        st.markdown("---")
        st.markdown(f"### Sube las fotos  ·  *{n} preguntas por prueba*")
        st.caption("Clave para que el motor OMR lea bien: las 4 columnas de RESPUESTAS visibles, foto "
                   "derecha, buena luz, sin fondo alrededor.")
        st.caption("💡 Procesa de a **15–20 fotos por lote** — si se corta la conexión, no pierdes el resto.")
        archivos = st.file_uploader("Fotos", type=["jpg","jpeg","png","webp"],
                                     accept_multiple_files=True, label_visibility="collapsed",
                                     key="uploader_fotos")

        # Cada foto seleccionada se guarda de inmediato en session_state (por hash de
        # contenido, no por nombre de archivo). Esto evita perder fotos anteriores cuando
        # el celular reabre la cámara y reemplaza la selección del input, y evita
        # colisiones cuando la cámara reutiliza el mismo nombre genérico (ej. "image.jpg").
        if archivos:
            hashes_conocidos = {v["hash"] for v in st.session_state.fotos_pendientes.values()}
            hashes_conocidos |= {d.get("hash") for d in st.session_state.resultados.values()}
            agregadas = 0
            for f in archivos:
                contenido = f.read()
                h = hashlib.md5(contenido).hexdigest()
                if h not in hashes_conocidos:
                    # Se guarda tal cual (sin comprimir aquí): el recorte en mitades y la
                    # compresión ocurren recién al procesar, a partir de la resolución
                    # original, para maximizar el detalle de cada acercamiento.
                    idx = len(st.session_state.fotos_pendientes) + len(st.session_state.resultados) + 1
                    st.session_state.fotos_pendientes[f"foto_{idx:03d}"] = {
                        "nombre": f.name, "bytes": contenido,
                        "mime": TIPOS_MIME.get(f.type, "image/jpeg"), "hash": h,
                    }
                    hashes_conocidos.add(h)
                    agregadas += 1
            if agregadas:
                st.rerun()

        # Mensaje del último procesamiento: se guarda en session_state porque el
        # st.rerun() de más abajo redibuja la pantalla de inmediato y borraría un
        # st.success/st.error mostrado en la misma corrida antes de que se alcance a leer.
        if st.session_state.get("mensaje_proceso"):
            tipo, texto = st.session_state.pop("mensaje_proceso")
            (st.error if tipo == "error" else st.success)(texto)

        pendientes = st.session_state.fotos_pendientes
        total_proc = len(st.session_state.resultados)

        if pendientes or total_proc:
            c1, c2 = st.columns(2)
            if total_proc: c1.success(f"✓ {total_proc} ya procesadas")
            if pendientes: c2.warning(f"⏳ {len(pendientes)} nuevas por procesar")

        if pendientes:
            with st.expander(f"📋 Fotos por procesar ({len(pendientes)})"):
                for id_unico, foto in pendientes.items():
                    st.caption(f"• {foto['nombre']}")
            key = api_key_activa()
            if not key:
                st.error("Ingresa tu API Key en el panel lateral.")
            elif st.button(f"🚀 Procesar {len(pendientes)} hoja(s)", type="primary", use_container_width=True):
                cliente = anthropic.Anthropic(api_key=key, timeout=90.0, max_retries=1)
                prog = st.progress(0, text="Iniciando...")
                errs = []
                items = list(pendientes.items())
                for i, (id_unico, foto) in enumerate(items):
                    prog.progress(i/len(items), text=f"Procesando {foto['nombre']} ({i+1}/{len(items)})...")
                    try:
                        solo_resp = (modo_actual == "solo_respuestas")
                        if st.session_state.get("usar_omr") and OMR_DISPONIBLE:
                            res = procesar_imagen_hibrido(cliente, foto["nombre"], foto["bytes"], foto["mime"], n,
                                                           solo_respuestas=solo_resp,
                                                           ia_arbitraje_habilitado=st.session_state.get(
                                                               "ia_arbitraje_habilitado", False))
                        else:
                            res = procesar_imagen(cliente, foto["nombre"], foto["bytes"], foto["mime"], n,
                                                   solo_respuestas=solo_resp)
                        res["hash"] = foto["hash"]
                        st.session_state.resultados[id_unico] = res
                        st.session_state.fotos_pendientes.pop(id_unico, None)
                    except anthropic.APITimeoutError:
                        errs.append(f"{foto['nombre']}: tiempo de espera agotado (conexión lenta) — quedó en la cola, reintenta.")
                    except Exception as e:
                        errs.append(f"{foto['nombre']}: {e} — quedó en la cola, reintenta.")
                prog.progress(1.0, text="¡Completado!")
                if errs:
                    st.session_state["mensaje_proceso"] = ("error", "Errores:\n"+"\n".join(errs))
                else:
                    st.session_state["mensaje_proceso"] = ("success", f"✅ {len(items)} procesadas. Ve a **Revisar y corregir**.")
                st.rerun()
        elif total_proc:
            st.success("✅ Todas las fotos cargadas. Ve a **Revisar y corregir**.")
        else:
            st.markdown("""
            <div style="border:2px dashed #d1d5db;border-radius:16px;padding:2.5rem;
                        text-align:center;color:#6b7280;margin-top:1rem;">
                <div style="font-size:3rem;">📷</div>
                <div style="font-size:16px;margin-top:10px;font-weight:500;">
                    Arrastra fotos aquí o toca para seleccionar</div>
                <div style="font-size:13px;margin-top:6px;">JPG · PNG · WEBP · Desde PC o celular</div>
            </div>""", unsafe_allow_html=True)


    # ══ REVISAR Y CORREGIR ═══════════════════════════════════════════════
    with tab_revisar:
        pauta = st.session_state["pauta"]

        if not st.session_state.resultados:
            st.info("Aún no hay hojas procesadas. Ve a **Cargar fotos**.")
        else:
            total_alumnos = len(st.session_state.resultados)
            pendientes = sum(
                1 for a,d in st.session_state.resultados.items()
                if [x for x in d.get("dudosas",[]) if str(x) not in st.session_state.correcciones.get(a,{})]
            )
            sospechosas = sum(1 for d in st.session_state.resultados.values() if d.get("sospechoso"))
            sin_id = sum(
                1 for a, d in st.session_state.resultados.items()
                if not (datos_efectivos(a).get("apellido_paterno") or datos_efectivos(a).get("nombres")
                         or datos_efectivos(a).get("cedula"))
            )
            m1,m2,m3,m4,m5 = st.columns(5)
            m1.metric("Alumnos procesados", total_alumnos)
            m2.metric("Con dudas pendientes", pendientes)
            m3.metric("Preguntas por prueba", n)
            m4.metric("🚨 Patrón sospechoso", sospechosas)
            m5.metric("🆔 Sin identificar", sin_id)
            if sospechosas:
                st.error(f"🚨 {sospechosas} hoja(s) con un patrón sospechoso o no leída — revísalas abajo antes de confiar en su puntaje.")
            if sin_id:
                st.warning(f"🆔 {sin_id} hoja(s) sin identificar — complétalas abajo antes de exportar.")
            st.markdown("---")

            for arch, datos_orig in st.session_state.resultados.items():
                datos     = datos_efectivos(arch)
                dudosas   = datos_orig.get("dudosas", [])
                corr_arch = st.session_state.correcciones.get(arch, {})
                pend_este = [d for d in dudosas if str(d) not in corr_arch]
                sk        = safe_key(arch)

                ref = respuestas_efectivas(arch)
                if pauta:
                    co,inc,om,_ = calcular(ref, pauta[:len(ref)])
                    total_p = sum(1 for p in pauta if p)
                    pct = round(co/total_p*100,1) if total_p else 0
                    puntaje_str = f"**{pct}%** ({co}/{total_p})"
                else:
                    puntaje_str = "_(sin pauta)_"

                es_sospechoso  = bool(datos_orig.get("sospechoso"))
                sin_identificar = not (datos.get("apellido_paterno") or datos.get("nombres") or datos.get("cedula"))
                icono = "🚨" if es_sospechoso else ("🆔" if sin_identificar else ("⚠️" if pend_este else "✅"))
                nombre_titulo = (f"{datos.get('apellido_paterno','')} {datos.get('nombres','')} — {datos.get('cedula','')}"
                                  if not sin_identificar
                                  else f"*(sin identificar — {datos_orig.get('archivo', arch)})*")
                with st.expander(
                    f"{icono} {nombre_titulo} | {puntaje_str}",
                    expanded=bool(pend_este or es_sospechoso or sin_identificar)
                ):
                    if es_sospechoso:
                        st.error(f"🚨 **Patrón sospechoso:** {datos_orig.get('motivo_sospecha','')}")
                    if sin_identificar:
                        st.warning("🆔 Falta identificar al alumno — completa al menos apellido paterno o cédula abajo.")

                    # ── Diagnóstico OMR ──────────────────────────────────
                    omr_meta = datos_orig.get("omr_meta")
                    if omr_meta and omr_meta.get("usado"):
                        n_total = len(ref)
                        pct_omr = round(omr_meta["n_confiable"] / n_total * 100) if n_total else 0
                        st.caption(f"🔬 **{pct_omr}%** confiable ({omr_meta['n_confiable']}/{n_total}) · "
                                   f"{omr_meta['n_dudosas']} para revisar abajo")
                        if omr_meta.get("n_geometry_error"):
                            bandas_bajas = [i + 1 for i, g in enumerate(omr_meta.get("geometry_confidence_por_banda", []))
                                            if g < OMR_THRESHOLDS["MIN_GEOMETRY_CONFIDENCE"]]
                            st.warning(f"🟣 {omr_meta['n_geometry_error']} pregunta(s) en la(s) columna(s) {bandas_bajas} "
                                       "no tienen evidencia de estar sobre burbujas reales (posible margen/texto vecino "
                                       "capturado de más) — quedaron para revisión manual, no se adivinaron.")
                        if datos_orig.get("omr_diagnostico_bytes"):
                            with st.expander("🔬 Ver diagnóstico OMR"):
                                st.caption("🟢 confiable · 🟡 confianza media · 🔵 resuelta por IA · 🟣 sin evidencia de grilla · 🔴 dudosa · ⚪ en blanco")
                                st.image(datos_orig["omr_diagnostico_bytes"], use_container_width=True)
                    elif omr_meta and not omr_meta.get("usado"):
                        st.error(f"🔬 No se pudo leer con OMR ({omr_meta.get('motivo_fallback','?')}). "
                                 "Completa a mano abajo o vuelve a tomar la foto con las 4 columnas visibles.")

                    # ── Datos editables del alumno ──────────────────────
                    if datos_orig.get("solo_respuestas"):
                        st.markdown("**Datos del alumno** *(modo solo-respuestas: complétalos a mano)*")
                    else:
                        st.markdown("**Datos del alumno** *(edita si Claude leyó mal algún campo)*")
                    ie = st.session_state.info_edits.get(arch, {})

                    r1c1, r1c2, r1c3 = st.columns(3)
                    with r1c1:
                        v = st.text_input("Apellido paterno",
                            value=ie.get("apellido_paterno", datos_orig.get("apellido_paterno","")),
                            key=f"ap_{sk}")
                        guardar_info_edit(arch, "apellido_paterno", v)
                    with r1c2:
                        v = st.text_input("Apellido materno",
                            value=ie.get("apellido_materno", datos_orig.get("apellido_materno","")),
                            key=f"am_{sk}")
                        guardar_info_edit(arch, "apellido_materno", v)
                    with r1c3:
                        v = st.text_input("Nombres",
                            value=ie.get("nombres", datos_orig.get("nombres","")),
                            key=f"no_{sk}")
                        guardar_info_edit(arch, "nombres", v)

                    r2c1, r2c2, r2c3 = st.columns(3)
                    with r2c1:
                        v = st.text_input("Cédula / RUT",
                            value=ie.get("cedula", datos_orig.get("cedula","")),
                            key=f"ce_{sk}")
                        guardar_info_edit(arch, "cedula", v)
                    with r2c2:
                        v = st.text_input("N° de folleto",
                            value=ie.get("nro_folleto", datos_orig.get("nro_folleto","")),
                            key=f"fo_{sk}")
                        guardar_info_edit(arch, "nro_folleto", v)
                    with r2c3:
                        st.markdown("&nbsp;")  # espaciador

                    # ── Respuestas dudosas ──────────────────────────────
                    if dudosas:
                        st.markdown("---")
                        st.markdown(f"**Preguntas dudosas:** {', '.join(f'P{d}' for d in dudosas)}")
                        crops_dudosas = datos_orig.get("omr_crops_dudosas", {})
                        if not crops_dudosas:
                            st.caption("Selecciona la respuesta correcta para cada una:")

                        # Renderizar en grupos de 6 (menos que antes: ahora cada columna
                        # también lleva el recorte de la fila, así que necesitan más ancho)
                        CHUNK = 6
                        for chunk_start in range(0, len(dudosas), CHUNK):
                            chunk = dudosas[chunk_start:chunk_start+CHUNK]
                            cols  = st.columns(len(chunk))
                            for j, num_p in enumerate(chunk):
                                resp_orig = (datos_orig["respuestas"][num_p-1]
                                             if num_p <= len(datos_orig["respuestas"]) else None)
                                # Leer valor guardado en session_state directamente para evitar blank
                                widget_key = f"dud_{sk}_{num_p}"
                                saved_val  = corr_arch.get(str(num_p), resp_orig or "—")
                                idx_opcion = OPCIONES.index(saved_val) if saved_val in OPCIONES else 5

                                with cols[j]:
                                    crop_bytes = crops_dudosas.get(str(num_p))
                                    if crop_bytes:
                                        st.image(crop_bytes, width=160)
                                    nueva = st.selectbox(
                                        f"P{num_p}",
                                        options=OPCIONES,
                                        index=idx_opcion,
                                        key=widget_key,
                                    )
                                    guardar_correccion(arch, num_p, nueva, resp_orig)
                    else:
                        st.markdown("---")
                        st.success("Sin preguntas dudosas.")

                    # ── Vista detalle de respuestas ─────────────────────
                    if pauta:
                        ref_act = respuestas_efectivas(arch)
                        st.markdown("---")
                        st.markdown("**Detalle de respuestas:**")
                        html="<div style='line-height:2.4;'>"
                        for i,(r,p) in enumerate(zip(ref_act,pauta),1):
                            es_dud  = i in dudosas
                            fue_c   = str(i) in st.session_state.correcciones.get(arch,{}) and es_dud
                            if r is None:
                                cls="badge-null"; txt=f"P{i}:—"
                            elif p and r.upper()==p.upper():
                                cls="badge-ok";  txt=f"P{i}:{r}"
                            else:
                                cls="badge-err"; txt=f"P{i}:{r or '?'}"
                            if es_dud and not fue_c: cls="badge-dud"
                            html+=f'<span class="corr-badge {cls}">{txt}{"✎" if fue_c else ""}</span>'
                        html+="</div>"
                        st.markdown(html, unsafe_allow_html=True)


    # ══ EXPORTAR ═════════════════════════════════════════════════════════
    with tab_exportar:
        pauta = st.session_state["pauta"]

        if not st.session_state.resultados:
            st.info("Procesa al menos una hoja antes de exportar.")
        elif not any(pauta):
            st.error("Ve a **Pauta de respuestas** e ingresa las respuestas correctas primero.")
        else:
            pend_exp = sum(
                1 for a,d in st.session_state.resultados.items()
                if [x for x in d.get("dudosas",[]) if str(x) not in st.session_state.correcciones.get(a,{})]
            )
            total_p=sum(1 for p in pauta if p)
            sospechosas_exp = sum(1 for d in st.session_state.resultados.values() if d.get("sospechoso"))
            sin_id_exp = sum(
                1 for a in st.session_state.resultados
                if not (datos_efectivos(a).get("apellido_paterno") or datos_efectivos(a).get("nombres")
                         or datos_efectivos(a).get("cedula"))
            )
            m1,m2,m3,m4,m5=st.columns(5)
            m1.metric("Alumnos",len(st.session_state.resultados))
            m2.metric("Preguntas evaluadas",total_p)
            m3.metric("Dudas pendientes",pend_exp,
                      delta="revisar antes" if pend_exp else "todo resuelto",
                      delta_color="inverse" if pend_exp else "normal")
            m4.metric("🚨 Sospechosas",sospechosas_exp,
                      delta="revisar antes" if sospechosas_exp else "ninguna",
                      delta_color="inverse" if sospechosas_exp else "normal")
            m5.metric("🆔 Sin identificar",sin_id_exp,
                      delta="completar antes" if sin_id_exp else "todo identificado",
                      delta_color="inverse" if sin_id_exp else "normal")

            if sospechosas_exp:
                st.error(f"🚨 {sospechosas_exp} hoja(s) sospechosa(s) o no leída(s) — revísalas en **Revisar y corregir**.")
            if sin_id_exp:
                st.warning(f"🆔 {sin_id_exp} hoja(s) sin identificar — quedarán en blanco en el Excel.")
            if pend_exp:
                st.warning(f"⚠️ {pend_exp} alumno(s) con dudas sin corregir — quedarán en amarillo.")
            elif not sospechosas_exp and not sin_id_exp:
                st.success("✅ Todo revisado. El Excel estará completo.")

            nombre_xl=f"resultados_{curso.replace(' ','_') if curso else 'curso'}.xlsx"
            excel=generar_excel(pauta, curso)
            st.download_button(
                label="⬇️ Descargar Excel", data=excel,
                file_name=nombre_xl,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True,
                # Respaldo de actividad (solo metadatos -- nunca nombres/RUT/
                # respuestas de alumnos, ver registro_sheets.py): se registra
                # justo cuando se entrega el resultado real, no antes.
                on_click=registro_sheets.registrar_evento,
                args=(st.session_state["auth_usuario"]["usuario"], "export_excel"),
                kwargs=dict(curso=curso, n_preguntas=total_p, n_alumnos=len(st.session_state.resultados),
                            dudas_pendientes=pend_exp, sospechosas=sospechosas_exp,
                            sin_identificar=sin_id_exp, omr_engine_version=OMR_ENGINE_VERSION),
            )
            st.caption("El Excel incluye Resumen, Detalle de respuestas, Estadísticas y Pauta utilizada — "
                       "la columna **Curso** en Resumen permite unir varios exports en un archivo maestro.")


# ═══════════════════════════════════════════════════════════════════════
# PLATAFORMA — página de inicio + router multipágina
# ═══════════════════════════════════════════════════════════════════════
def render_inicio():
    """Página de bienvenida a la plataforma: presenta las herramientas
    disponibles (hoy, solo el Revisor de Pruebas) y deja espacio visual para
    las que se agreguen después."""
    st.markdown("# 🎓 Herramientas Docentes")
    st.caption("Herramientas para apoyar la revisión y corrección de recursos educativos.")
    st.markdown("---")

    with st.container(border=True):
        col_desc, col_boton = st.columns([3, 1])
        with col_desc:
            st.markdown("### 📝 Revisor de Hojas de Respuestas")
            st.markdown(
                "Corrige hojas de respuestas tipo PAES (80 preguntas, 4 bloques, "
                "alternativas A-E) a partir de fotos: detecta la grilla automáticamente, "
                "lee las marcas y arma el Excel de resultados."
            )
            st.markdown('<span class="corr-badge badge-ok">Disponible</span>', unsafe_allow_html=True)
        with col_boton:
            st.markdown("")
            st.page_link(pagina_revisor, label="Abrir →", icon="📝", use_container_width=True)

    st.caption("✨ Más herramientas de IA para docentes están en camino.")


# Convención para agregar una herramienta nueva a la plataforma:
#   1. Una función render_xxx() que dibuje esa herramienta -- en este mismo
#      archivo si es chica y comparte código con el motor OMR, o en un
#      archivo nuevo importado acá si es grande e independiente.
#   2. Una entrada st.Page(render_xxx, title=..., icon=...) en el diccionario
#      de abajo, dentro de "Herramientas".
#   3. Una tarjeta correspondiente en render_inicio().
st.set_page_config(
    page_title="Herramientas Docentes",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded",
)


def _render_login():
    """Formulario de acceso -- grupo cerrado, cuentas precargadas por el
    administrador (ver gestionar_usuarios.py); no hay auto-registro. Bloquea
    TODA la plataforma (Inicio y cualquier herramienta) hasta iniciar
    sesión con una cuenta válida y activa."""
    st.markdown("# 🎓 Herramientas Docentes")
    st.caption("Acceso restringido — ingresa con tu cuenta.")
    with st.form("form_login"):
        usuario = st.text_input("Usuario")
        password = st.text_input("Clave", type="password")
        enviado = st.form_submit_button("Ingresar", type="primary", use_container_width=True)
    if enviado:
        datos = registro_sheets.verificar_login(usuario, password)
        if datos:
            st.session_state["auth_usuario"] = datos
            registro_sheets.registrar_evento(datos["usuario"], "login")
            st.rerun()
        else:
            st.error("Usuario o clave incorrectos.")


if not st.session_state.get("auth_usuario"):
    _render_login()
    st.stop()

with st.sidebar:
    _auth = st.session_state["auth_usuario"]
    st.markdown(f"👤 **{_auth['nombre']}**")
    if st.button("Cerrar sesión", use_container_width=True):
        del st.session_state["auth_usuario"]
        st.rerun()
    st.markdown("---")

pagina_inicio = st.Page(render_inicio, title="Inicio", icon="🏠", default=True, url_path="inicio")
pagina_revisor = st.Page(render_revisor_pruebas, title="Revisor de Pruebas", icon="📝",
                          url_path="revisor-de-pruebas")

pg = st.navigation({
    "Plataforma": [pagina_inicio],
    "Herramientas": [pagina_revisor],
})
pg.run()
