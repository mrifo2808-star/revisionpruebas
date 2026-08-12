"""
Tests de la carga de pauta desde Excel (alternativa de ingreso de las
respuestas correctas, además del grid y el pegado rápido "Importar rápido"):
`generar_plantilla_pauta_excel` produce una plantilla de 2 columnas (N°,
Respuesta) con validación de datos A-E; `pauta_desde_excel_bytes` la lee de
vuelta (o cualquier Excel razonable con una columna de respuesta
reconocible) y devuelve una pauta de largo n más avisos legibles -- nunca
una excepción, un archivo mal formado se trata como "sin respuestas".

Estas dos funciones son independientes de Streamlit (no llaman a st.*), así
que se prueban extrayéndolas del código fuente de app_revisor.py con la
misma técnica que usa omr_metrics._cargar_app_module() para el motor OMR,
sin necesidad de ejecutar la UI completa.

Correr desde la raíz del repo:  py -m pytest tests/test_pauta_excel.py -q
"""
import io
import os
import re
import sys

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, RAIZ)


def _extraer(src, nombre):
    start = src.index(f"\ndef {nombre}") + 1
    rest = src[start + len(f"def {nombre}"):]
    m = re.search(r'\ndef |\n\n\n#', rest)
    end = start + len(f"def {nombre}") + (m.start() if m else len(rest))
    return src[start:end]


@pytest.fixture(scope="session")
def funcs():
    src = open(os.path.join(RAIZ, "app_revisor.py"), encoding="utf-8").read()
    ns = {"io": io, "pd": pd, "Workbook": Workbook, "Font": Font,
          "PatternFill": PatternFill, "Alignment": Alignment,
          "DataValidation": DataValidation, "LETRAS_VALIDAS": {"A", "B", "C", "D", "E"}}
    exec(_extraer(src, "generar_plantilla_pauta_excel"), ns)
    exec(_extraer(src, "pauta_desde_excel_bytes"), ns)
    return ns


def _wb_con_respuestas(letras, header=("N°", "Respuesta")):
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, header[0])
    ws.cell(1, 2, header[1])
    for i, l in enumerate(letras):
        ws.cell(i + 2, 1, f"P{i + 1}")
        ws.cell(i + 2, 2, l)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_plantilla_vacia_no_produce_respuestas(funcs):
    print("=== la plantilla recien descargada (sin llenar) no debe interpretarse como respuestas ===")
    n = 10
    plantilla = funcs["generar_plantilla_pauta_excel"](n)
    assert len(plantilla) > 0
    pauta, avisos = funcs["pauta_desde_excel_bytes"](plantilla, n)
    assert pauta == [None] * n
    assert avisos == []
    print("OK\n")


def test_roundtrip_plantilla_llena(funcs):
    print("=== plantilla generada por la app, llenada a mano, se lee de vuelta exacto ===")
    n = 10
    plantilla = funcs["generar_plantilla_pauta_excel"](n)
    wb = load_workbook(io.BytesIO(plantilla))
    ws = wb.active
    letras = ["A", "B", "C", "D", "E", "A", "B", "C", "D", "E"]
    for i, l in enumerate(letras):
        ws.cell(i + 2, 2, l)
    buf = io.BytesIO()
    wb.save(buf)
    pauta, avisos = funcs["pauta_desde_excel_bytes"](buf.getvalue(), n)
    assert pauta == letras
    assert avisos == []
    print("OK\n")


def test_valores_invalidos_quedan_en_blanco_no_crashea(funcs):
    print("=== un valor que no es A-E se deja en blanco (nunca inventa una letra), y se avisa ===")
    n = 5
    datos = _wb_con_respuestas(["A", "X", "c", "", "Z"])
    pauta, avisos = funcs["pauta_desde_excel_bytes"](datos, n)
    assert pauta == ["A", None, "C", None, None]
    assert any("2 valor" in a for a in avisos)
    print("OK\n")


def test_menos_filas_que_n_completa_con_none(funcs):
    print("=== archivo con menos filas que preguntas configuradas -- se completa con None, no crashea ===")
    n = 8
    datos = _wb_con_respuestas(["A", "B", "C"])
    pauta, avisos = funcs["pauta_desde_excel_bytes"](datos, n)
    assert pauta == ["A", "B", "C", None, None, None, None, None]
    assert any("3 fila" in a and "8 preguntas" in a for a in avisos)
    print("OK\n")


def test_mas_filas_que_n_se_recorta(funcs):
    print("=== archivo con mas filas que preguntas configuradas -- se recorta a n, no se pierde nada critico ===")
    n = 3
    datos = _wb_con_respuestas(["A", "B", "C", "D", "E"])
    pauta, avisos = funcs["pauta_desde_excel_bytes"](datos, n)
    assert pauta == ["A", "B", "C"]
    assert any("5 fila" in a for a in avisos)
    print("OK\n")


def test_reconoce_nombres_de_columna_alternativos(funcs):
    print("=== columnas con otro nombre razonable (Alternativa, Correcta, Pauta, Clave) tambien se reconocen ===")
    n = 4
    for nombre_col in ("Alternativa", "Correcta", "Pauta", "Clave", "alternativa"):
        datos = _wb_con_respuestas(["A", "B", "C", "D"], header=("Pregunta", nombre_col))
        pauta, avisos = funcs["pauta_desde_excel_bytes"](datos, n)
        assert pauta == ["A", "B", "C", "D"], f"fallo con columna '{nombre_col}': {pauta}"
    print("OK\n")


def test_archivo_vacio_no_crashea(funcs):
    print("=== un Excel sin filas de datos no debe lanzar excepcion, solo devolver una pauta vacia con aviso ===")
    n = 5
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "N°"); ws.cell(1, 2, "Respuesta")
    buf = io.BytesIO()
    wb.save(buf)
    pauta, avisos = funcs["pauta_desde_excel_bytes"](buf.getvalue(), n)
    assert pauta == [None] * n
    print("OK\n")


def test_archivo_no_excel_no_crashea(funcs):
    print("=== bytes que no son un Excel valido -- debe devolver pauta vacia + aviso, nunca una excepcion ===")
    n = 5
    pauta, avisos = funcs["pauta_desde_excel_bytes"](b"esto no es un excel", n)
    assert pauta == [None] * n
    assert avisos, "deberia avisar que no se pudo leer el archivo"
    print("OK\n")


def test_plantilla_tiene_validacion_de_datos_a_e(funcs):
    print("=== la plantilla generada trae una lista desplegable A-E en la columna Respuesta ===")
    n = 6
    plantilla = funcs["generar_plantilla_pauta_excel"](n)
    wb = load_workbook(io.BytesIO(plantilla))
    ws = wb.active
    dvs = list(ws.data_validations.dataValidation)
    assert len(dvs) == 1
    assert dvs[0].formula1 == '"A,B,C,D,E"'
    assert ws.cell(1, 2).value == "Respuesta"
    assert ws.cell(2, 1).value == "P1"
    assert ws.cell(n + 1, 1).value == f"P{n}"
    print("OK\n")


if __name__ == "__main__":
    import omr_metrics  # solo para reusar el mismo patron de ejecucion standalone
    src = open(os.path.join(RAIZ, "app_revisor.py"), encoding="utf-8").read()
    ns = {"io": io, "pd": pd, "Workbook": Workbook, "Font": Font,
          "PatternFill": PatternFill, "Alignment": Alignment,
          "DataValidation": DataValidation, "LETRAS_VALIDAS": {"A", "B", "C", "D", "E"}}
    exec(_extraer(src, "generar_plantilla_pauta_excel"), ns)
    exec(_extraer(src, "pauta_desde_excel_bytes"), ns)
    test_plantilla_vacia_no_produce_respuestas(ns)
    test_roundtrip_plantilla_llena(ns)
    test_valores_invalidos_quedan_en_blanco_no_crashea(ns)
    test_menos_filas_que_n_completa_con_none(ns)
    test_mas_filas_que_n_se_recorta(ns)
    test_reconoce_nombres_de_columna_alternativos(ns)
    test_archivo_vacio_no_crashea(ns)
    test_archivo_no_excel_no_crashea(ns)
    test_plantilla_tiene_validacion_de_datos_a_e(ns)
    print("TODO PASO - pauta desde excel OK")
